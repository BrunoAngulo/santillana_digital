#!/usr/bin/env python3
"""
Crea módulos y recursos en el LMS de Santillana a partir de la estructura local.

Flujo por producto:
  1. Detecta carpetas bajo PRIM/ que tienen OUTPUT/resumen_global.xlsx
  2. Lee el Excel: A=Nombre visible, B=Nombre archivo, C=Ruta, D=Tipo,
                   E=Carpeta contenedora (módulo), F=Carpeta fuente, G=Name (sección)
  3. Aplica fix de nombres:
       - "Unidad XX" → "Name Unidad XX"   (col A empieza con "Unidad")
       - nombre == stem archivo → "Name nombre"  (col A igual al stem de col B)
  4. Busca el curso en el LMS (búsqueda interactiva)
  5. Por cada módulo del Excel (orden: Documentos → Unidad 00 → Unidad 01 …):
       a. Crea una lección (módulo) si no existe ya
       b. Por cada sección dentro del módulo (col G, orden original del Excel):
            - Crea la sección con el nombre de col G
            - Busca cada archivo en CMS por ERP ID y lo agrega a esa sección

Uso:
  python crear_programa.py
  python crear_programa.py "C:\\...\\SUBIDAS\\PRIM"
"""

import re
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from tqdm import tqdm
except ImportError:
    print("Instalando dependencias...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "requests", "openpyxl", "tqdm"])
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from tqdm import tqdm


API_BASE = "https://compartirconocimientos-pe.santillana.com"

CONTENT_TYPES = [
    "CTTY_04", "CTTY_01", "CTTY_08", "CTTY_07", "CTTY_06",
    "CTTY_12", "CTTY_05", "CTTY_03", "CTTY_20", "CTTY_13",
]

CARPETAS_EXCLUIR = {"completo"}  # carpetas que no se convierten en módulo

# Mapeo carpeta fuente (col F) → nombre de sección visible en el LMS
SECCION_NOMBRE = {
    "dif_lect":  "Dificultades de lectoescritura",
    "doc_curr":  "Documentos curriculares",
    "est_lect":  "Estrategias de lectura",
    "guia_met":  "Guía metodológica",
    "lam_didac": "Láminas didácticas",
    "libmed":    "Libromedia",
    "mat_imp":   "Módulo de material imprimible",
    "mod_ie":    "Instrumentos de evaluación",
    "senpai":    "Herramientas cooperativas",
}

# Orden de secciones dentro de cada módulo (Unidades)
_ORDEN_SECCIONES = [
    "Documentos curriculares",
    "Libromedia",
    "Estrategias de lectura",
    "Dificultades de lectoescritura",
    "Guía metodológica",
    "Instrumentos de evaluación",
    "Módulo de material imprimible",
    "Láminas didácticas",
    "Herramientas cooperativas",
    "Recursos",
]

# Orden de módulos en el curso (los derivados de col E = Documentos van primero)
_ORDEN_MODULOS_CURSO = [
    "Documentos",
    "Documentos curriculares",
    "Libromedia",
    "Estrategias de lectura",
    "Dificultades de lectoescritura",
    "Guía metodológica",
    "Instrumentos de evaluación",
    "Módulo de material imprimible",
    "Láminas didácticas",
    "Herramientas cooperativas",
    # Unidades van al final, ordenadas numéricamente por reordenar_modulos_curso
]


# ── Sesión HTTP ──────────────────────────────────────────────────────────────

def build_session(token: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "Authorization": token.strip(),
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Origin":   "https://publisher.compartirconocimientos-pe.santillana.com",
        "Referer":  "https://publisher.compartirconocimientos-pe.santillana.com/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/149.0.0.0 Safari/537.36 Edg/149.0.0.0"
        ),
    })
    return s


_RETRY_STATUS = {502, 503, 504}
_RETRY_DELAYS = [3, 6, 12]   # segundos entre reintentos


def _request_with_retry(fn, *args, **kwargs):
    """Ejecuta fn(*args, **kwargs) reintentando hasta 3 veces en errores 5xx transitorios."""
    for intento, delay in enumerate([0] + _RETRY_DELAYS, 1):
        if delay:
            time.sleep(delay)
        try:
            r = fn(*args, **kwargs)
            if r.status_code in _RETRY_STATUS:
                tqdm.write(f"  [RETRY {intento}/4] {r.status_code} en {r.url} — reintentando en {_RETRY_DELAYS[intento - 1]}s...")
                continue
            r.raise_for_status()
            return r
        except requests.ConnectionError as e:
            if intento <= len(_RETRY_DELAYS):
                tqdm.write(f"  [RETRY {intento}/4] ConnectionError — reintentando...")
                continue
            raise
    r.raise_for_status()
    return r


def api_get(session, path: str, params=None):
    r = _request_with_retry(session.get, f"{API_BASE}{path}", params=params, timeout=30)
    return r.json()


def api_post(session, path: str, payload: dict):
    r = _request_with_retry(session.post, f"{API_BASE}{path}", json=payload, timeout=30)
    return r.json()


# ── Detección de productos ───────────────────────────────────────────────────

def detectar_productos(base_path: Path) -> list[Path]:
    """Devuelve subcarpetas de base_path que contienen OUTPUT/resumen_global.xlsx."""
    return sorted(
        p for p in base_path.iterdir()
        if p.is_dir() and (p / "OUTPUT" / "resumen_global.xlsx").exists()
    )


# ── Lectura y normalización del Excel ────────────────────────────────────────

def _fix_nombre(nombre_visible: str, nombre_archivo: str, name_serie: str) -> str:
    """
    Si nombre_visible empieza con "Unidad" o es igual al stem del archivo,
    antepone el nombre de la serie: "Dificultades de lectoescritura Unidad 01".
    """
    v     = nombre_visible.strip()
    stem  = Path(nombre_archivo).stem if nombre_archivo else ""
    serie = name_serie.strip() if name_serie else ""

    if serie and (v.lower().startswith("unidad") or v == stem):
        return f"{serie} {v}"
    return v


def leer_excel_producto(xlsx_path: Path) -> tuple[list[dict], bool]:
    """
    Devuelve (archivos, tiene_carpeta_docs).
    Filas con col E == "Documentos": modulo = col F mapeado, seccion = "Recursos".
    Resto: modulo = col E (Unidad 01…), seccion = col F mapeado.
    tiene_carpeta_docs indica si hay que crear el módulo "Documentos" vacío.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    archivos = []
    tiene_carpeta_docs = False

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        cols = list(row) + [None] * 8
        nombre_visible = str(cols[0] or "").strip()
        nombre_archivo = str(cols[1] or "").strip()
        carpeta_cont   = str(cols[4] or "").strip()   # col E: Unidad 01 / Documentos…
        carpeta_fuente = str(cols[5] or "").strip()   # col F: dif_lect, doc_curr…
        name_serie     = str(cols[6] or "").strip()   # col G: prefijo para nombres

        if not nombre_archivo or not carpeta_cont:
            continue

        erp_id       = Path(nombre_archivo).stem
        nombre_final = _fix_nombre(nombre_visible, nombre_archivo, name_serie)

        if carpeta_cont.lower() == "documentos":
            tiene_carpeta_docs = True
            modulo  = SECCION_NOMBRE.get(carpeta_fuente, carpeta_fuente)
            seccion = "Recursos"
        else:
            modulo  = carpeta_cont
            seccion = SECCION_NOMBRE.get(carpeta_fuente, carpeta_fuente)

        archivos.append({
            "nombre_visible": nombre_final,
            "erp_id":         erp_id,
            "modulo":         modulo,
            "seccion":        seccion,
        })

    return archivos, tiene_carpeta_docs


def agrupar_por_modulo(archivos: list[dict]) -> dict[str, dict[str, list[dict]]]:
    """
    Devuelve {modulo → {seccion → [archivos]}}.
    Excluye módulos en CARPETAS_EXCLUIR.
    Orden módulos: SECCION_NOMBRE-derivados primero (filas Documentos),
                   luego Unidades numéricamente.
    Orden secciones dentro de cada módulo: según posición en SECCION_NOMBRE.
    """
    estructura: dict[str, dict[str, list[dict]]] = {}
    secciones_orden: dict[str, list[str]] = {}

    for a in archivos:
        modulo  = a["modulo"]
        seccion = a["seccion"]

        if modulo.lower() in CARPETAS_EXCLUIR:
            continue

        if modulo not in estructura:
            estructura[modulo] = {}
            secciones_orden[modulo] = []

        if seccion not in estructura[modulo]:
            estructura[modulo][seccion] = []
            secciones_orden[modulo].append(seccion)

        estructura[modulo][seccion].append(a)

    def sort_key_modulo(nombre: str) -> tuple:
        try:
            return (0, _ORDEN_MODULOS_CURSO.index(nombre), "")
        except ValueError:
            m = re.search(r'\d+', nombre)
            return (1, int(m.group()) if m else 9999, nombre.lower())

    def sort_key_seccion(nombre: str) -> tuple:
        try:
            return (0, _ORDEN_SECCIONES.index(nombre))
        except ValueError:
            return (1, nombre.lower())

    resultado: dict[str, dict[str, list[dict]]] = {}
    for modulo in sorted(estructura.keys(), key=sort_key_modulo):
        resultado[modulo] = {
            sec: estructura[modulo][sec]
            for sec in sorted(secciones_orden[modulo], key=sort_key_seccion)
        }

    return resultado


# ── API: cursos ───────────────────────────────────────────────────────────────

def buscar_cursos(session, search: str) -> list[dict]:
    data = api_get(session, "/api/lms/courses", params={
        "offset": 0, "page": 0, "pageSize": 20,
        "isEditorial": 1, "search": search,
    })
    return data.get("data", {}).get("courses", [])


def get_course_items(session, course_guid: str) -> list[dict]:
    data = api_get(session, f"/api/front/courses/{course_guid}/items")
    return data.get("data", {}).get("items", [])


def crear_lesson(session, course_guid: str, nombre_modulo: str) -> str | None:
    """Crea una lección (módulo) en el curso. Devuelve su lesson_guid."""
    html_name = f'<p><span style="font-size: 24px;">{nombre_modulo}</span></p>'
    payload = {
        "evaluation_period_id": "annual",
        "name":               html_name,
        "number_of_sessions": 1,
        "type":               "teacher",
    }
    data = api_post(session, f"/api/front/courses/{course_guid}/lessons", payload)
    d = data.get("data", {})
    return d.get("lesson_guid") or d.get("guid")


def crear_seccion(session, lesson_guid: str, nombre_seccion: str) -> str | None:
    """Crea una sección dentro de la lección. Devuelve su guid (parent_guid)."""
    nombre_seccion = nombre_seccion.strip() or "Recursos"
    payload = {
        "lesson_guid":       lesson_guid,
        "section":           nombre_seccion,
        "section_type_guid": "default",
    }
    r = session.post(f"{API_BASE}/api/front/lesson-items", json=payload, timeout=30)
    try:
        r.raise_for_status()
    except requests.HTTPError as e:
        tqdm.write(f"      [ERROR HTTP] crear_seccion: {e} — {r.text[:200]}")
        return None
    data = r.json()
    guid = (data.get("data") or {}).get("guid")
    if not guid:
        tqdm.write(f"      [WARN] crear_seccion sin guid — resp: {str(data)[:200]}")
    return guid


# ── API: contenidos CMS ───────────────────────────────────────────────────────

def buscar_contenido_por_erp(session, erp_id: str) -> dict | None:
    params = [
        ("offset", 0), ("page", 0), ("pageSize", 5),
        ("search", erp_id), ("isEditorial", 1),
    ]
    for t in CONTENT_TYPES:
        params.append(("type[]", t))

    r = _request_with_retry(session.get, f"{API_BASE}/api/cms/contents",
                            params=params, timeout=30)
    contents = r.json().get("data", {}).get("contents", [])

    # Coincidencia exacta primero
    for c in contents:
        if c.get("erp_id", "").lower() == erp_id.lower():
            return c
    return contents[0] if contents else None


def agregar_contenido(session, lesson_guid: str, parent_guid: str,
                      content: dict, nombre_visible: str) -> str | None:
    """Agrega el contenido a la sección. Devuelve el item_guid creado o None si falla."""
    payload = {
        "content_guid":       content["guid"],
        "description":        "",
        "disciplines":        [d["discipline_guid"]        for d in content.get("disciplines",     [])],
        "educationLevels":    [e["education_level_guid"]   for e in content.get("educationLevels",  [])],
        "educationYears":     [a["education_year_guid"]    for a in content.get("educationYears",   [])],
        "include_in_gradebook": 0,
        "is_embed":           1,
        "item_for":           "all",
        "learningObjectives": [],
        "lesson_guid":        lesson_guid,
        "name":               nombre_visible,
        "parent_guid":        parent_guid,
        "status":             "published",
        "teacher_notes":      None,
    }
    data = api_post(session, "/api/front/lesson-items", payload)
    return data.get("data", {}).get("guid") or None


# ── API: secciones existentes y reordenamiento ───────────────────────────────

def get_lesson_sections(session, lesson_guid: str) -> dict[str, str]:
    """Devuelve {nombre_seccion_lower → guid} para las secciones ya existentes."""
    try:
        data  = api_get(session, f"/api/front/lessons/{lesson_guid}/items")
        items = (data.get("data") or {}).get("items", [])
        return {
            (it.get("section") or "").strip().lower(): it["guid"]
            for it in items
            if it.get("section") and it.get("guid")
        }
    except Exception:
        return {}


def reordenar_secciones(session, lesson_guid: str) -> bool:
    """
    Reordena las secciones del módulo según _ORDEN_SECCIONES.
    Se llama siempre — para módulos nuevos y existentes.
    """
    try:
        data  = api_get(session, f"/api/front/lessons/{lesson_guid}/items")
        items = (data.get("data") or {}).get("items", [])
        secciones = [
            ((it.get("section") or it.get("name") or "").strip(), it["guid"])
            for it in items if it.get("guid")
        ]
        if not secciones:
            return True

        def _key(nom: str):
            try:
                return (0, _ORDEN_SECCIONES.index(nom))
            except ValueError:
                return (1, nom.lower())

        secciones_ord = sorted(secciones, key=lambda x: _key(x[0]))
        reorder = [{"guid": g, "order": i + 1}
                   for i, (_, g) in enumerate(secciones_ord)]
        r = session.put(f"{API_BASE}/api/front/lesson-items",
                        json={"reorder": reorder}, timeout=30)
        r.raise_for_status()
        return r.json().get("status") == "success"
    except Exception as e:
        tqdm.write(f"      [ERROR] reordenar_secciones: {e}")
        return False


def reordenar_modulos_curso(session, course_guid: str) -> bool:
    """
    Reordena los módulos del curso según _ORDEN_MODULOS_CURSO;
    los no reconocidos (Unidades) van al final ordenados numéricamente.
    """
    try:
        items = get_course_items(session, course_guid)
        modulos = [
            (html_a_texto(it.get("lesson_name") or it.get("name") or ""), it["guid"])
            for it in items if it.get("guid")
        ]
        if not modulos:
            return True

        def _key(nom: str):
            try:
                return (0, _ORDEN_MODULOS_CURSO.index(nom), "")
            except ValueError:
                m = re.search(r'\d+', nom)
                return (1, int(m.group()) if m else 9999, nom.lower())

        modulos_ord = sorted(modulos, key=lambda x: _key(x[0]))
        reorder = [{"guid": g, "order": i + 1}
                   for i, (_, g) in enumerate(modulos_ord)]
        r = session.put(f"{API_BASE}/api/front/courses/{course_guid}/items",
                        json={"reorder": reorder}, timeout=30)
        r.raise_for_status()
        ok = r.json().get("status") == "success"
        if ok:
            tqdm.write(f"  [OK] Módulos del curso reordenados ({len(reorder)} módulos)")
        else:
            tqdm.write(f"  [WARN] reordenar_modulos: {r.text[:120]}")
        return ok
    except Exception as e:
        tqdm.write(f"  [ERROR] reordenar_modulos_curso: {e}")
        return False


# ── Log Excel ────────────────────────────────────────────────────────────────

_FILL_HDR   = PatternFill("solid", fgColor="1F4E79")
_FONT_HDR   = Font(bold=True, color="FFFFFF", size=11)
_FILL_OK    = PatternFill("solid", fgColor="C6EFCE")
_FILL_ERR   = PatternFill("solid", fgColor="FFC7CE")
_FILL_MISS  = PatternFill("solid", fgColor="FFEB9C")   # amarillo: NOT FOUND
_FILL_SKIP  = PatternFill("solid", fgColor="D9D9D9")


def guardar_log(entradas: list[dict], log_path: Path):
    """
    Genera un Excel con una fila por cada contenido procesado.
    Columnas: Producto | Módulo | Sección | ERP ID | Nombre visible |
              Estado | GUID CMS | Nombre en CMS | GUID item LMS | Detalle
    Colores: VERDE=CREADO, ROJO=ERROR, AMARILLO=NOT_FOUND, GRIS=SKIP
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Log"

    headers = [
        "Producto", "Módulo", "Sección", "ERP ID", "Nombre visible",
        "Estado", "GUID CMS", "Nombre en CMS", "GUID item LMS", "Detalle",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _FILL_HDR
        c.font = _FONT_HDR
        c.alignment = Alignment(horizontal="center")

    for i, e in enumerate(entradas, 2):
        ws.cell(row=i, column=1,  value=e.get("producto",       ""))
        ws.cell(row=i, column=2,  value=e.get("modulo",         ""))
        ws.cell(row=i, column=3,  value=e.get("seccion",        ""))
        ws.cell(row=i, column=4,  value=e.get("erp_id",         ""))
        ws.cell(row=i, column=5,  value=e.get("nombre_visible", ""))
        ws.cell(row=i, column=6,  value=e.get("estado",         ""))
        ws.cell(row=i, column=7,  value=e.get("guid_cms",       ""))
        ws.cell(row=i, column=8,  value=e.get("nombre_cms",     ""))
        ws.cell(row=i, column=9,  value=e.get("guid_item_lms",  ""))
        ws.cell(row=i, column=10, value=e.get("detalle",        ""))

        estado = e.get("estado", "")
        if estado == "CREADO":
            fill = _FILL_OK
        elif estado == "NOT_FOUND":
            fill = _FILL_MISS
        elif estado == "SKIP":
            fill = _FILL_SKIP
        else:
            fill = _FILL_ERR

        for col in range(1, 11):
            ws.cell(row=i, column=col).fill = fill

    anchos = {"A": 25, "B": 30, "C": 35, "D": 30, "E": 45,
              "F": 12, "G": 38, "H": 45, "I": 38, "J": 40}
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    wb.save(log_path)
    print(f"  Log guardado: {log_path.resolve()}")


# ── Selección interactiva ─────────────────────────────────────────────────────

def seleccionar(titulo: str, opciones: list[str]) -> int:
    print(f"\n  {titulo}:")
    for i, op in enumerate(opciones, 1):
        print(f"    {i}. {op}")
    while True:
        try:
            idx = int(input(f"  Selección (1-{len(opciones)}) > ").strip())
            if 1 <= idx <= len(opciones):
                return idx - 1
        except (ValueError, EOFError):
            pass
        print("  Opción inválida.")


# ── Extracción de texto de HTML ───────────────────────────────────────────────

def html_a_texto(html: str) -> str:
    return re.sub(r'<[^>]+>', '', html or "").strip()


# ── Procesamiento de un producto ──────────────────────────────────────────────

def procesar_producto(session, producto_path: Path) -> dict:
    xlsx = producto_path / "OUTPUT" / "resumen_global.xlsx"
    print(f"\n{'─' * 62}")
    print(f"  Producto : {producto_path.name}")

    archivos, tiene_carpeta_docs = leer_excel_producto(xlsx)
    estructura = agrupar_por_modulo(archivos)
    modulos    = list(estructura.keys())

    print(f"  Módulos  : {modulos}")

    # Buscar curso
    print()
    search_term = input(f"  Búsqueda del curso para '{producto_path.name}' > ").strip()
    cursos = buscar_cursos(session, search_term)

    if not cursos:
        print("  [AVISO] No se encontraron cursos. Saltando.")
        return {"producto": producto_path.name, "error": "sin curso", "entradas_log": []}

    if len(cursos) == 1:
        curso = cursos[0]
        print(f"  Curso: {curso['name']}  [{curso['guid']}]")
    else:
        idx = seleccionar(
            "Selecciona el curso",
            [f"{c['name']} — {c.get('education_year_name','?')} — {c.get('discipline_name','?')}"
             for c in cursos],
        )
        curso = cursos[idx]

    course_guid = curso["guid"]

    # Módulos ya existentes (para no duplicar)
    items_existentes   = get_course_items(session, course_guid)
    modulos_existentes = {html_a_texto(it.get("lesson_name", "")).lower(): it
                          for it in items_existentes}

    # Crear módulo "Documentos" vacío si el Excel tiene filas con col E = Documentos
    if tiene_carpeta_docs:
        if "documentos" not in modulos_existentes:
            doc_guid = crear_lesson(session, course_guid, "Documentos")
            if doc_guid:
                tqdm.write(f"    [+] Módulo vacío: Documentos  [{doc_guid}]")
            else:
                tqdm.write("    [ERROR] No se pudo crear el módulo 'Documentos'")
        else:
            tqdm.write("    [SKIP] Módulo 'Documentos' ya existe")

    modulos_creados    = 0
    modulos_salteados  = 0
    secciones_creadas  = 0
    contenidos_ok      = 0
    contenidos_error   = 0
    entradas_log: list[dict] = []

    def _log(modulo, seccion, archivo, estado, content=None, item_guid=None, detalle=""):
        entradas_log.append({
            "producto":       producto_path.name,
            "modulo":         modulo,
            "seccion":        seccion,
            "erp_id":         archivo.get("erp_id", "") if archivo else "",
            "nombre_visible": archivo.get("nombre_visible", "") if archivo else "",
            "estado":         estado,
            "guid_cms":       content["guid"]  if content else "",
            "nombre_cms":     content.get("name", "") if content else "",
            "guid_item_lms":  item_guid or "",
            "detalle":        detalle,
        })

    barra = tqdm(modulos, desc=f"  {producto_path.name}", unit="módulo", ncols=80)

    for modulo in barra:
        barra.set_postfix_str(modulo[:30])

        if modulo.lower() in modulos_existentes:
            lesson_guid = modulos_existentes[modulo.lower()].get("lesson_guid", "")
            tqdm.write(f"    [SKIP] Módulo ya existe: {modulo}  [{lesson_guid}]")
            modulos_salteados += 1
            for seccion, archivos_sec in estructura[modulo].items():
                for archivo in archivos_sec:
                    _log(modulo, seccion, archivo, "SKIP", detalle="Módulo ya existía")
            # Aún así verificar el orden de las secciones
            if lesson_guid:
                reordenar_secciones(session, lesson_guid)
            continue

        # Crear módulo
        lesson_guid = crear_lesson(session, course_guid, modulo)
        if not lesson_guid:
            tqdm.write(f"    [ERROR] No se pudo crear módulo: {modulo}")
            for seccion, archivos_sec in estructura[modulo].items():
                for archivo in archivos_sec:
                    _log(modulo, seccion, archivo, "ERROR", detalle="Fallo al crear módulo")
            continue
        tqdm.write(f"    [+] Módulo: {modulo}  [{lesson_guid}]")
        modulos_creados += 1

        # Por cada sección dentro del módulo
        secciones_existentes = get_lesson_sections(session, lesson_guid)
        for seccion, archivos_sec in estructura[modulo].items():
            clave = seccion.strip().lower()
            if clave in secciones_existentes:
                parent_guid = secciones_existentes[clave]
                tqdm.write(f"      [SKIP] Sección ya existe: {seccion}  [{parent_guid}]")
            else:
                parent_guid = crear_seccion(session, lesson_guid, seccion)
                if not parent_guid:
                    tqdm.write(f"      [ERROR] No se pudo crear sección '{seccion}'")
                    for archivo in archivos_sec:
                        _log(modulo, seccion, archivo, "ERROR",
                             detalle=f"Fallo al crear sección '{seccion}'")
                    continue
                tqdm.write(f"      [+] Sección: {seccion}  [{parent_guid}]")
                secciones_creadas += 1

            # Agregar contenidos a la sección
            for archivo in archivos_sec:
                content = buscar_contenido_por_erp(session, archivo["erp_id"])
                if not content:
                    tqdm.write(f"        [NOT FOUND] {archivo['erp_id']}")
                    contenidos_error += 1
                    _log(modulo, seccion, archivo, "NOT_FOUND",
                         detalle="ERP ID no encontrado en CMS")
                else:
                    item_guid = agregar_contenido(
                        session, lesson_guid, parent_guid,
                        content, archivo["nombre_visible"],
                    )
                    if item_guid:
                        tqdm.write(f"        [+] {archivo['nombre_visible']}")
                        contenidos_ok += 1
                        _log(modulo, seccion, archivo, "CREADO",
                             content=content, item_guid=item_guid)
                    else:
                        tqdm.write(f"        [ERROR] {archivo['erp_id']}")
                        contenidos_error += 1
                        _log(modulo, seccion, archivo, "ERROR",
                             content=content, detalle="PUT lesson-item falló")

                time.sleep(0.15)   # evitar rate-limit

        # Verificar orden de secciones dentro de este módulo
        reordenar_secciones(session, lesson_guid)

    # Verificar orden de módulos en el curso
    print(f"  Verificando orden de módulos del curso...")
    reordenar_modulos_curso(session, course_guid)

    return {
        "producto":          producto_path.name,
        "curso":             curso["name"],
        "modulos_creados":   modulos_creados,
        "modulos_salteados": modulos_salteados,
        "secciones_creadas": secciones_creadas,
        "contenidos_ok":     contenidos_ok,
        "contenidos_error":  contenidos_error,
        "entradas_log":      entradas_log,
    }


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  Creador de programa LMS - Santillana Digital")
    print("=" * 62)
    print()

    if len(sys.argv) > 1:
        base_str = sys.argv[1].strip().strip('"').strip("'")
    else:
        base_str = input(
            "Ruta base PRIM (ej: C:\\...\\SUBIDAS\\PRIM) > "
        ).strip().strip('"').strip("'")

    base_path = Path(base_str)
    if not base_path.is_dir():
        sys.exit(f"\n[ERROR] La carpeta no existe: {base_path}")

    token = input("\nToken de autorización (Bearer ...) > ").strip()
    if not token:
        sys.exit("No se ingresó token.")

    session   = build_session(token)
    productos = detectar_productos(base_path)

    if not productos:
        sys.exit("No se encontraron carpetas con OUTPUT/resumen_global.xlsx.")

    print(f"\nProductos detectados: {len(productos)}")
    for i, p in enumerate(productos, 1):
        print(f"  {i}. {p.name}")

    selec = input("\n¿Procesar todos? [s] o números separados por comas > ").strip().lower()
    if selec not in ("", "s"):
        nums = [int(x.strip()) - 1 for x in selec.split(",") if x.strip().isdigit()]
        productos = [productos[i] for i in nums if 0 <= i < len(productos)]

    resultados = [procesar_producto(session, p) for p in productos]

    # ── Log Excel ────────────────────────────────────────────────────
    todas_entradas = []
    for r in resultados:
        todas_entradas.extend(r.get("entradas_log", []))

    if todas_entradas:
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = base_path / f"log_programa_{ts}.xlsx"
        print()
        guardar_log(todas_entradas, log_path)

    # ── Resumen final ────────────────────────────────────────────────
    print(f"\n{'=' * 62}")
    print("  RESUMEN FINAL")
    print("=" * 62)
    for r in resultados:
        if "error" in r:
            print(f"  {r['producto']:<20} ERROR: {r['error']}")
        else:
            print(f"  {r['producto']}")
            print(f"    Curso             : {r['curso']}")
            print(f"    Módulos creados   : {r['modulos_creados']}")
            print(f"    Módulos saltados  : {r['modulos_salteados']}")
            print(f"    Secciones creadas : {r['secciones_creadas']}")
            print(f"    Contenidos OK     : {r['contenidos_ok']}")
            print(f"    Contenidos error  : {r['contenidos_error']}")
    print("=" * 62)


if __name__ == "__main__":
    main()
