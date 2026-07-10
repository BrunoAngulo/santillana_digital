#!/usr/bin/env python3
"""
Gestiona los contenidos vinculados a un programa (curso) del LMS.

Modos:
  SIN argumento  → busca el curso, lista todos sus items y genera un Excel
  CON Excel      → lee el Excel editado y actualiza los nombres que cambiaron

Columna editable : "Nombre visible"
Columnas fijas   : Módulo, Sección, ERP ID, GUID item LMS, GUID CMS

Uso:
  python gestionar_programa.py
  python gestionar_programa.py items_MiCurso_20260710.xlsx
"""

import re
import sys
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from tqdm import tqdm
except ImportError:
    print("Instalando dependencias...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "requests", "openpyxl", "tqdm"])
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from tqdm import tqdm


API_BASE    = "https://compartirconocimientos-pe.santillana.com"
MAX_WORKERS = 5


# ── Sesión ────────────────────────────────────────────────────────────────────

def build_session(token: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "Authorization": token.strip(),
        "Content-Type":  "application/json",
        "Accept":        "application/json, text/plain, */*",
        "Origin":  "https://publisher.compartirconocimientos-pe.santillana.com",
        "Referer": "https://publisher.compartirconocimientos-pe.santillana.com/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/149.0.0.0 Safari/537.36 Edg/149.0.0.0"
        ),
    })
    return s


# ── Helpers ───────────────────────────────────────────────────────────────────

def html_a_texto(html: str) -> str:
    return re.sub(r"<[^>]+>", "", html or "").strip()


def buscar_cursos(session, search: str) -> list[dict]:
    r = session.get(f"{API_BASE}/api/lms/courses", params={
        "offset": 0, "page": 0, "pageSize": 20,
        "isEditorial": 1, "search": search,
    }, timeout=30)
    r.raise_for_status()
    return r.json().get("data", {}).get("courses", [])


def seleccionar(titulo: str, opciones: list[str]) -> int:
    print(f"\n  {titulo}:")
    for i, op in enumerate(opciones, 1):
        print(f"    {i}. {op}")
    while True:
        try:
            idx = int(input(f"  Selección (1-{len(opciones)}) > ").strip())
            if 1 <= idx <= len(opciones):
                return idx - 1
        except (ValueError, EOFError):
            pass
        print("  Opción inválida.")


# ── GET items del curso ───────────────────────────────────────────────────────

def get_course_items(session, course_guid: str) -> list[dict]:
    r = session.get(f"{API_BASE}/api/front/courses/{course_guid}/items",
                    timeout=30)
    r.raise_for_status()
    return r.json().get("data", {}).get("items", [])


def get_lesson_items(session, lesson_guid: str) -> list[dict]:
    """Devuelve los items (secciones + contenidos) dentro de una lección."""
    r = session.get(f"{API_BASE}/api/front/lessons/{lesson_guid}/items",
                    timeout=30)
    if r.status_code == 404:
        return []
    r.raise_for_status()
    data = r.json().get("data", {})
    # La respuesta puede venir como lista o como dict con clave "items"
    if isinstance(data, list):
        return data
    return data.get("items", [])


def _fetch_lesson_items(token: str, lesson: dict) -> tuple[dict, list[dict]]:
    """Thread-safe: obtiene los items de una lección."""
    s = build_session(token)
    items = get_lesson_items(s, lesson["lesson_guid"])
    return lesson, items


# ── Parseo de la jerarquía ────────────────────────────────────────────────────

def construir_filas(token: str, course_items: list[dict]) -> list[dict]:
    """
    Recorre lecciones → secciones → contenidos y devuelve una fila por contenido.
    Cada fila: modulo, seccion, nombre_visible, erp_id, guid_item, guid_cms, content_data
    """
    # Separar lecciones (tienen lesson_guid y lesson_name)
    lecciones = []
    seen = set()
    for it in course_items:
        lg = it.get("lesson_guid") or it.get("guid")
        if lg and lg not in seen:
            seen.add(lg)
            lecciones.append({
                "lesson_guid": lg,
                "lesson_name": html_a_texto(it.get("lesson_name") or it.get("name") or ""),
            })

    print(f"  {len(lecciones)} módulos encontrados. Obteniendo items...")

    # Obtener items de cada lección en paralelo
    lesson_items_map: dict[str, list[dict]] = {}
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_fetch_lesson_items, token, lec): lec
                   for lec in lecciones}
        done = 0
        for future in as_completed(futures):
            lec, items = future.result()
            lesson_items_map[lec["lesson_guid"]] = items
            done += 1
            print(f"  {done}/{len(lecciones)} módulos obtenidos")

    # Construir filas aplanadas
    filas = []
    for lec in lecciones:
        modulo = lec["lesson_name"]
        items  = lesson_items_map.get(lec["lesson_guid"], [])

        # Identificar secciones (items sin content_guid) y contenidos (con content_guid)
        secciones: dict[str, str] = {}   # guid_seccion → nombre_seccion
        contenidos = []

        for it in items:
            content_guid = (it.get("content_guid") or
                            it.get("content", {}).get("guid") if isinstance(it.get("content"), dict) else None)
            if not content_guid:
                # Es una sección contenedora
                guid_sec = it.get("guid", "")
                nom_sec  = (it.get("section") or it.get("name") or "")
                if guid_sec:
                    secciones[guid_sec] = nom_sec
            else:
                contenidos.append(it)

        for it in contenidos:
            parent_guid  = it.get("parent_guid", "")
            seccion_nom  = secciones.get(parent_guid, "")
            content_obj  = it.get("content") if isinstance(it.get("content"), dict) else {}
            erp_id       = (content_obj.get("erp_id") or
                            it.get("erp_id") or "")
            guid_cms     = (content_obj.get("guid") or
                            it.get("content_guid") or "")

            filas.append({
                "modulo":         modulo,
                "seccion":        seccion_nom,
                "nombre_visible": it.get("name", ""),
                "erp_id":         erp_id,
                "guid_item":      it.get("guid", ""),
                "guid_cms":       guid_cms,
            })

    return filas


# ── Excel: exportar ───────────────────────────────────────────────────────────

_FILL_HDR  = PatternFill("solid", fgColor="1F4E79")
_FONT_HDR  = Font(bold=True, color="FFFFFF", size=11)
_FILL_PAR  = PatternFill("solid", fgColor="DDEEFF")
_FILL_IMPAR= PatternFill("solid", fgColor="FFFFFF")
_FILL_LOCK = PatternFill("solid", fgColor="F2F2F2")   # columnas no editables


HEADERS = [
    "Módulo", "Sección", "Nombre visible",
    "ERP ID", "GUID item LMS", "GUID CMS",
]
COL_EDITABLE = "Nombre visible"   # única columna que el usuario debe editar


def guardar_excel_lista(filas: list[dict], xlsx_path: Path, curso_nombre: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    # Cabecera
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _FILL_HDR
        c.font = _FONT_HDR
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 18

    # Filas
    for i, fila in enumerate(filas, 2):
        fill_base = _FILL_PAR if i % 2 == 0 else _FILL_IMPAR
        vals = [
            fila["modulo"],
            fila["seccion"],
            fila["nombre_visible"],
            fila["erp_id"],
            fila["guid_item"],
            fila["guid_cms"],
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            # Columnas no editables con fondo diferente
            c.fill = fill_base if HEADERS[col - 1] == COL_EDITABLE else _FILL_LOCK

    # Anchos
    for letra, ancho in zip("ABCDEF", [35, 35, 50, 30, 38, 38]):
        ws.column_dimensions[letra].width = ancho

    # Nota en la hoja
    ws["A1"].comment = None
    ws.sheet_properties.tabColor = "1F4E79"

    # Hoja oculta con todos los GUIDs originales (para detectar eliminaciones)
    wm = wb.create_sheet("_meta")
    wm["A1"] = "curso"
    wm["B1"] = curso_nombre
    wm["A2"] = "guids_originales"
    for i, fila in enumerate(filas, 3):
        wm.cell(row=i, column=1, value=fila["guid_item"])
        wm.cell(row=i, column=2, value=fila["modulo"])
        wm.cell(row=i, column=3, value=fila["seccion"])
        wm.cell(row=i, column=4, value=fila["nombre_visible"])
    wm.sheet_state = "hidden"

    wb.save(xlsx_path)
    print(f"  Excel guardado: {xlsx_path.resolve()}")


# ── Excel: leer para actualizar ───────────────────────────────────────────────

def leer_excel_ediciones(xlsx_path: Path) -> tuple[list[dict], list[dict]]:
    """
    Devuelve (filas_actuales, filas_eliminadas).
    filas_actuales  → filas que siguen en la hoja Items (para actualizar nombre)
    filas_eliminadas → filas que estaban en _meta pero ya no están en Items (para DELETE)
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Items"]

    headers = [str(c.value or "").strip() for c in ws[1]]
    try:
        col_guid  = headers.index("GUID item LMS")
        col_nom   = headers.index("Nombre visible")
        col_mod   = headers.index("Módulo")
        col_sec   = headers.index("Sección")
    except ValueError as e:
        sys.exit(f"\n[ERROR] Columna no encontrada: {e}")

    filas_actuales = []
    guids_actuales: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        guid = str(row[col_guid] or "").strip()
        nom  = str(row[col_nom]  or "").strip()
        if guid and nom:
            filas_actuales.append({
                "guid_item":    guid,
                "nombre_nuevo": nom,
                "modulo":       str(row[col_mod] or "").strip(),
                "seccion":      str(row[col_sec] or "").strip(),
            })
            guids_actuales.add(guid)

    # Leer GUIDs originales de la hoja _meta
    filas_eliminadas = []
    if "_meta" in wb.sheetnames:
        wm = wb["_meta"]
        for row in wm.iter_rows(min_row=3, values_only=True):
            guid = str(row[0] or "").strip()
            if guid and guid not in guids_actuales:
                filas_eliminadas.append({
                    "guid_item": guid,
                    "modulo":    str(row[1] or "").strip(),
                    "seccion":   str(row[2] or "").strip(),
                    "nombre":    str(row[3] or "").strip(),
                })

    return filas_actuales, filas_eliminadas


# ── GET / PUT item individual ─────────────────────────────────────────────────

def get_lesson_item(session, item_guid: str) -> dict | None:
    try:
        r = session.get(f"{API_BASE}/api/front/lesson-items/{item_guid}",
                        timeout=30)
        r.raise_for_status()
        return r.json().get("data")
    except requests.RequestException as e:
        tqdm.write(f"  [ERROR GET] {item_guid}: {e}")
        return None


def actualizar_nombre_item(session, item_guid: str, existente: dict,
                           nombre_nuevo: str) -> bool:
    """
    PUT /api/front/lesson-items/{guid} con el nombre nuevo.
    Mantiene todos los demás campos del item actual.
    """
    payload = {k: v for k, v in existente.items()}
    payload["name"] = nombre_nuevo

    try:
        r = session.put(f"{API_BASE}/api/front/lesson-items/{item_guid}",
                        json=payload, timeout=30)
        r.raise_for_status()
        resp = r.json()
        return resp.get("status") == "success" or bool(resp.get("data"))
    except requests.RequestException as e:
        tqdm.write(f"  [ERROR PUT] {item_guid}: {e}")
        return False


# ── DELETE items ─────────────────────────────────────────────────────────────

def eliminar_items(session, guids: list[str]) -> tuple[list[str], list[str]]:
    """
    DELETE /api/front/lesson-items con {"guid": [...]}.
    Devuelve (eliminados, fallidos).
    """
    if not guids:
        return [], []
    try:
        r = session.delete(
            f"{API_BASE}/api/front/lesson-items",
            json={"guid": guids},
            timeout=30,
        )
        r.raise_for_status()
        data = r.json()
        if data.get("status") == "success":
            return guids, []
        return [], guids
    except requests.RequestException as e:
        tqdm.write(f"  [ERROR DELETE] {e}")
        return [], guids


# ── Worker actualización ──────────────────────────────────────────────────────

def _procesar_actualizacion(token: str, fila: dict) -> dict:
    session      = build_session(token)
    guid         = fila["guid_item"]
    nombre_nuevo = fila["nombre_nuevo"]
    resultado    = {
        "guid_item":      guid,
        "modulo":         fila["modulo"],
        "seccion":        fila["seccion"],
        "nombre_nuevo":   nombre_nuevo,
        "nombre_actual":  "",
        "estado":         "ERROR",
        "detalle":        "",
    }

    existente = get_lesson_item(session, guid)
    if existente is None:
        resultado["detalle"] = "No se pudo obtener el item"
        return resultado

    nombre_actual = existente.get("name", "")
    resultado["nombre_actual"] = nombre_actual

    if nombre_actual == nombre_nuevo:
        resultado["estado"]  = "SIN_CAMBIO"
        resultado["detalle"] = "Nombre idéntico"
        return resultado

    ok = actualizar_nombre_item(session, guid, existente, nombre_nuevo)
    if ok:
        resultado["estado"]  = "ACTUALIZADO"
        resultado["detalle"] = f'"{nombre_actual}" → "{nombre_nuevo}"'
    else:
        resultado["detalle"] = "PUT falló"

    return resultado


# ── Log actualización ─────────────────────────────────────────────────────────

def guardar_log_actualizacion(resultados: list[dict], log_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cambios"

    headers = ["GUID item LMS", "Módulo", "Sección",
               "Nombre actual", "Nombre nuevo", "Estado", "Detalle"]
    fill_h   = PatternFill("solid", fgColor="1F4E79")
    font_h   = Font(bold=True, color="FFFFFF", size=11)
    fill_ok  = PatternFill("solid", fgColor="C6EFCE")
    fill_del = PatternFill("solid", fgColor="F4CCCC")   # rojo claro: eliminado
    fill_err = PatternFill("solid", fgColor="FFC7CE")
    fill_nc  = PatternFill("solid", fgColor="D9D9D9")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_h
        c.font = font_h
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(resultados, 2):
        ws.cell(row=i, column=1, value=r["guid_item"])
        ws.cell(row=i, column=2, value=r.get("modulo", ""))
        ws.cell(row=i, column=3, value=r.get("seccion", ""))
        ws.cell(row=i, column=4, value=r.get("nombre_actual", ""))
        ws.cell(row=i, column=5, value=r.get("nombre_nuevo", ""))
        ws.cell(row=i, column=6, value=r["estado"])
        ws.cell(row=i, column=7, value=r.get("detalle", ""))

        estado = r["estado"]
        fill = (fill_ok  if estado == "ACTUALIZADO" else
                fill_del if estado == "ELIMINADO"   else
                fill_nc  if estado == "SIN_CAMBIO"  else
                fill_err)
        for col in range(1, 8):
            ws.cell(row=i, column=col).fill = fill

    for letra, ancho in zip("ABCDEFG", [38, 30, 30, 50, 50, 14, 55]):
        ws.column_dimensions[letra].width = ancho

    wb.save(log_path)
    print(f"  Log guardado: {log_path.resolve()}")


# ── Modo 1: Listar ────────────────────────────────────────────────────────────

def modo_listar(token: str):
    session = build_session(token)

    search = input("Búsqueda del curso > ").strip()
    cursos = buscar_cursos(session, search)
    if not cursos:
        sys.exit("No se encontraron cursos.")

    if len(cursos) == 1:
        curso = cursos[0]
        print(f"  Curso: {curso['name']}  [{curso['guid']}]")
    else:
        idx = seleccionar(
            "Selecciona el curso",
            [f"{c['name']} — {c.get('education_year_name','?')} — {c.get('discipline_name','?')}"
             for c in cursos],
        )
        curso = cursos[idx]

    course_guid  = curso["guid"]
    curso_nombre = curso["name"]

    print(f"\nObteniendo estructura del curso...")
    course_items = get_course_items(session, course_guid)

    if not course_items:
        sys.exit("El curso no tiene items.")

    filas = construir_filas(token, course_items)

    if not filas:
        sys.exit("No se encontraron contenidos vinculados.")

    print(f"\n  {len(filas)} contenidos encontrados.")

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_nom = re.sub(r'[^\w\-]', '_', curso_nombre)[:40]
    xlsx_path = Path(f"items_{safe_nom}_{ts}.xlsx")

    guardar_excel_lista(filas, xlsx_path, curso_nombre)

    print(f"\n{'=' * 62}")
    print(f"  Contenidos : {len(filas)}")
    print(f"  Excel      : {xlsx_path.resolve()}")
    print(f"\n  Edita la columna 'Nombre visible' y ejecuta:")
    print(f"  python gestionar_programa.py \"{xlsx_path.name}\"")
    print("=" * 62)


# ── Modo 2: Actualizar ────────────────────────────────────────────────────────

def modo_actualizar(token: str, xlsx_path: Path):
    print(f"Leyendo {xlsx_path.name}...")
    filas, filas_eliminar = leer_excel_ediciones(xlsx_path)
    print(f"  {len(filas)} filas presentes  |  {len(filas_eliminar)} filas eliminadas")

    if not filas and not filas_eliminar:
        sys.exit("No hay filas válidas.")

    session    = build_session(token)
    resultados = []

    # ── Eliminaciones (batch) ────────────────────────────────────────
    if filas_eliminar:
        print(f"\nEliminando {len(filas_eliminar)} items del programa...")
        guids_del = [f["guid_item"] for f in filas_eliminar]
        ok_guids, err_guids = eliminar_items(session, guids_del)
        ok_set  = set(ok_guids)

        for f in filas_eliminar:
            if f["guid_item"] in ok_set:
                print(f"  [ELIMINADO] {f['modulo']} / {f['seccion']} / {f['nombre']}")
                resultados.append({
                    "guid_item":    f["guid_item"],
                    "modulo":       f["modulo"],
                    "seccion":      f["seccion"],
                    "nombre_actual": f["nombre"],
                    "nombre_nuevo": "",
                    "estado":       "ELIMINADO",
                    "detalle":      "Fila eliminada del Excel",
                })
            else:
                print(f"  [ERROR] No se pudo eliminar: {f['guid_item']}")
                resultados.append({
                    "guid_item":    f["guid_item"],
                    "modulo":       f["modulo"],
                    "seccion":      f["seccion"],
                    "nombre_actual": f["nombre"],
                    "nombre_nuevo": "",
                    "estado":       "ERROR",
                    "detalle":      "DELETE falló",
                })

    # ── Actualizaciones de nombre (paralelo) ─────────────────────────
    if filas:
        print(f"\nActualizando nombres (máximo {MAX_WORKERS} simultáneos)...\n")
        barra = tqdm(total=len(filas), desc="Actualizando", unit="item", ncols=80)

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
            futures = {pool.submit(_procesar_actualizacion, token, f): f for f in filas}
            for future in as_completed(futures):
                r = future.result()
                resultados.append(r)
                if r["estado"] == "ACTUALIZADO":
                    tqdm.write(f"  [OK] {r['detalle']}")
                elif r["estado"] == "SIN_CAMBIO":
                    tqdm.write(f"  [=]  {r['nombre_actual']}")
                else:
                    tqdm.write(f"  [ERROR] {r['guid_item']}: {r['detalle']}")
                barra.update(1)

        barra.close()

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = xlsx_path.parent / f"log_gestion_{ts}.xlsx"
    print()
    guardar_log_actualizacion(resultados, log_path)

    act = sum(1 for r in resultados if r["estado"] == "ACTUALIZADO")
    eli = sum(1 for r in resultados if r["estado"] == "ELIMINADO")
    nc  = sum(1 for r in resultados if r["estado"] == "SIN_CAMBIO")
    err = sum(1 for r in resultados if r["estado"] == "ERROR")

    print(f"\n{'=' * 62}")
    print(f"  Actualizados  : {act}")
    print(f"  Eliminados    : {eli}")
    print(f"  Sin cambio    : {nc}")
    print(f"  Errores       : {err}")
    print(f"  Log           : {log_path}")
    print("=" * 62)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  Gestión de programa LMS - Santillana Digital")
    print("=" * 62)
    print()

    token = input("Token de autorización (Bearer ...) > ").strip()
    if not token:
        sys.exit("No se ingresó token.")
    print()

    if len(sys.argv) > 1:
        xlsx_str  = sys.argv[1].strip().strip('"').strip("'")
        xlsx_path = Path(xlsx_str)
        if not xlsx_path.exists():
            sys.exit(f"[ERROR] Archivo no encontrado: {xlsx_path}")
        modo_actualizar(token, xlsx_path)
    else:
        modo_listar(token)


if __name__ == "__main__":
    main()
