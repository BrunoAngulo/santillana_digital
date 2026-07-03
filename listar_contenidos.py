#!/usr/bin/env python3
"""
Lista contenidos creados desde una fecha en Compartir Conocimientos (Santillana).
Pagina /api/cms/contents (ordenado por created_at desc) y se detiene al llegar
a items más antiguos que la fecha indicada.
Para cada contenido obtiene el detalle (versions) y detecta el tipo real del
archivo adjunto. Genera un Excel con columna "Tipo archivo" coloreada:
  - VERDE  → tipo coincide con "Tipo contenido"
  - ROJO   → tipo no coincide
  - GRIS   → sin archivo adjunto

Uso:
  python listar_contenidos.py
  python listar_contenidos.py 2026-06-25
"""

import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timezone
from pathlib import Path

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Instalando dependencias...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests", "openpyxl"])
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment


API_BASE     = "https://compartirconocimientos-pe.santillana.com"
CONTENTS_URL = f"{API_BASE}/api/cms/contents"
PAGE_SIZE    = 100
DETAIL_WORKERS = 10   # peticiones paralelas al endpoint de detalle

DISP_LABEL = {0: "Docentes y Estudiantes", 1: "Docentes"}

# Extensión del archivo en versions[] → tipo que devuelve la API (type_name)
_EXT_TIPO = {
    ".pdf":  "PDF",
    ".html": "HTML Interactivo",
    ".htm":  "HTML Interactivo",
    ".zip":  "HTML Interactivo",
    ".docx": "OFFICE",
    ".doc":  "OFFICE",
    ".pptx": "OFFICE",
    ".xlsx": "OFFICE",
}


# ── Sesión ────────────────────────────────────────────────────────────────────

def build_session(token: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "Authorization": token.strip(),
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Origin":   "https://publisher.compartirconocimientos-pe.santillana.com",
        "Referer":  "https://publisher.compartirconocimientos-pe.santillana.com/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/149.0.0.0 Safari/537.36 Edg/149.0.0.0"
        ),
    })
    return s


# ── Paginación ────────────────────────────────────────────────────────────────

def get_contents_desde(session: requests.Session, desde: date) -> list[dict]:
    """
    Devuelve todos los contenidos con created_at >= desde.
    El endpoint devuelve items ordenados por created_at desc, así que en
    cuanto aparece uno anterior a la fecha de corte se puede parar.
    """
    desde_dt = datetime(desde.year, desde.month, desde.day, tzinfo=timezone.utc)
    result   = []
    offset   = 0

    while True:
        params = {
            "offset":      offset,
            "page":        offset // PAGE_SIZE,
            "orderBy":     "created_at desc",
            "pageSize":    PAGE_SIZE,
            "author":      "editorial",
            "search":      "",
            "isEditorial": 1,
        }
        resp = session.get(CONTENTS_URL, params=params, timeout=30)
        resp.raise_for_status()
        body = resp.json()

        if body.get("status") != "success":
            print(f"  [ERROR] respuesta inesperada: {body}")
            break

        data     = body["data"]
        total    = data["total"]
        contents = data["contents"]

        if not contents:
            break

        stop = False
        for c in contents:
            created_at = datetime.fromisoformat(c["created_at"].replace("Z", "+00:00"))
            if created_at < desde_dt:
                stop = True
                break
            result.append(c)

        print(f"  offset={offset:>5} | página {offset // PAGE_SIZE + 1} | "
              f"acumulados {len(result):>4} / {total}")

        offset += PAGE_SIZE
        if stop or offset >= total:
            break

    return result


# ── Detalle por contenido (versions) ─────────────────────────────────────────

def _fetch_version_url(token: str, guid: str) -> tuple[str, str]:
    """
    Llama a GET /api/cms/contents/{guid} y devuelve (guid, url_del_archivo).
    La URL viene de versions[-1].url (versión más reciente).
    Crea su propia sesión para ser thread-safe.
    """
    s = build_session(token)
    try:
        r = s.get(f"{API_BASE}/api/cms/contents/{guid}", timeout=30)
        r.raise_for_status()
        d = r.json().get("data", {})
        versions = d.get("versions", [])
        if versions:
            # La API devuelve versions en orden ascendente; la última = más reciente
            return guid, versions[-1].get("url", "")
        # Fallback: url del campo principal
        return guid, d.get("url", "")
    except Exception:
        return guid, ""


def obtener_version_urls(token: str, contents: list[dict]) -> dict[str, str]:
    """Descarga en paralelo la URL de versión de cada contenido."""
    result: dict[str, str] = {}
    total = len(contents)
    print(f"\nObteniendo detalle de {total} contenidos (paralelo)...")

    with ThreadPoolExecutor(max_workers=DETAIL_WORKERS) as pool:
        futures = {pool.submit(_fetch_version_url, token, c["guid"]): c["guid"]
                   for c in contents}
        done = 0
        for future in as_completed(futures):
            guid, url = future.result()
            result[guid] = url
            done += 1
            if done % 10 == 0 or done == total:
                print(f"  {done}/{total} detalles obtenidos")

    return result


# ── Tipo de archivo desde URL ─────────────────────────────────────────────────

def tipo_desde_url(url: str) -> str:
    """Detecta el tipo de contenido a partir de la extensión del archivo en la URL."""
    if not url:
        return ""
    ext = Path(url.split("?")[0]).suffix.lower()
    return _EXT_TIPO.get(ext, "")


# ── Extracción de campos ──────────────────────────────────────────────────────

def extraer_fila(c: dict, version_url: str = "") -> dict:
    url_path       = c.get("url", "")
    nombre_archivo = Path(url_path).name if url_path else ""

    colecciones = ", ".join(col["collection"]            for col in c.get("collections",    []))
    etapas      = ", ".join(e["education_level_name"]    for e   in c.get("educationLevels", []))
    anios       = ", ".join(a["education_year_name"]     for a   in c.get("educationYears",  []))
    asignaturas = ", ".join(d["discipline_name"]         for d   in c.get("disciplines",     []))
    idiomas     = ", ".join(l["name"]                    for l   in c.get("langs",            []))

    disp         = DISP_LABEL.get(c.get("is_teacher_only", 0), "Desconocido")
    tipo_archivo = tipo_desde_url(version_url) if version_url else ""

    return {
        "Nombre archivo":   nombre_archivo,
        "Colección":        colecciones,
        "Etapa":            etapas,
        "Año / serie":      anios,
        "Asignaturas":      asignaturas,
        "Idioma":           idiomas,
        "Nombre contenido": c.get("name", ""),
        "Tipo contenido":   c.get("type_name", ""),
        "Tipo archivo":     tipo_archivo,      # verde/rojo/gris en Excel
        "Disponible para":  disp,
        "ERP":              c.get("erp_id", ""),
        "Creado":           c.get("created_at", "")[:10],
        "GUID":             c.get("guid", ""),
    }


# ── Excel ─────────────────────────────────────────────────────────────────────

_FILL_VERDE       = PatternFill("solid", fgColor="C6EFCE")   # coincide
_FILL_ROJO        = PatternFill("solid", fgColor="FFC7CE")   # no coincide
_FILL_GRIS        = PatternFill("solid", fgColor="D9D9D9")   # sin archivo
_FILL_SIN_ARCHIVO_TXT = "Sin archivo"


def guardar_excel(filas: list[dict], xlsx_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contenidos"

    if not filas:
        wb.save(xlsx_path)
        return

    headers   = list(filas[0].keys())
    fill_h    = PatternFill("solid", fgColor="1F4E79")
    font_h    = Font(bold=True, color="FFFFFF", size=11)
    fill_par  = PatternFill("solid", fgColor="DDEEFF")
    fill_impar= PatternFill("solid", fgColor="FFFFFF")

    # Índice de las columnas clave (1-based)
    col_tipo_archivo   = headers.index("Tipo archivo")   + 1
    col_tipo_contenido = headers.index("Tipo contenido") + 1

    # Cabecera
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill_h
        cell.font = font_h
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    # Filas
    for i, fila in enumerate(filas, 2):
        fill_base = fill_par if i % 2 == 0 else fill_impar
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=i, column=col, value=fila[key])
            cell.fill = fill_base

        # Color de "Tipo archivo"
        tipo_arch = (fila.get("Tipo archivo") or "").strip()
        tipo_cont = (fila.get("Tipo contenido") or "").strip()
        cell_arch = ws.cell(row=i, column=col_tipo_archivo)

        if not tipo_arch:
            cell_arch.value = _FILL_SIN_ARCHIVO_TXT
            cell_arch.fill  = _FILL_GRIS
        elif tipo_arch.lower() == tipo_cont.lower():
            cell_arch.fill = _FILL_VERDE
        else:
            cell_arch.fill = _FILL_ROJO

    # Anchos
    anchos = {
        "A": 35, "B": 40, "C": 20, "D": 35, "E": 20,
        "F": 15, "G": 50, "H": 20, "I": 20, "J": 25,
        "K": 30, "L": 12, "M": 38,
    }
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    wb.save(xlsx_path)
    print(f"  Guardado: {xlsx_path.resolve()}")


# ── Función reutilizable ──────────────────────────────────────────────────────

def listar_contenidos_desde(session: requests.Session, token: str,
                             desde: date) -> list[dict]:
    """Devuelve lista de dicts con los campos clave de cada contenido."""
    raw          = get_contents_desde(session, desde)
    version_urls = obtener_version_urls(token, raw)
    return [extraer_fila(c, version_urls.get(c["guid"], "")) for c in raw]


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  Listar contenidos por fecha - Santillana Digital")
    print("=" * 62)
    print()

    if len(sys.argv) > 1:
        desde_str = sys.argv[1]
    else:
        desde_str = input("Desde fecha (YYYY-MM-DD) [2026-06-25]: ").strip() or "2026-06-25"

    try:
        desde = date.fromisoformat(desde_str)
    except ValueError:
        sys.exit(f"[ERROR] Fecha inválida: {desde_str}  (formato esperado: YYYY-MM-DD)")

    print()
    token = input("Token de autorización (Bearer ...) > ").strip()
    if not token:
        sys.exit("No se ingresó token.")

    print(f"\nBuscando contenidos creados desde {desde} ...\n")
    session = build_session(token)
    filas   = listar_contenidos_desde(session, token, desde)

    if not filas:
        print(f"\nNo se encontraron contenidos desde {desde}.")
        return

    # Resumen en consola
    print()
    print(f"{'#':<5} {'Nombre archivo':<36} {'Tipo API':<20} {'Tipo archivo':<20} {'ERP'}")
    print("─" * 110)
    for i, f in enumerate(filas, 1):
        tipo_a = f["Tipo archivo"] or "Sin archivo"
        ok     = "✓" if tipo_a.lower() == f["Tipo contenido"].lower() else "✗"
        print(f"{i:<5} {f['Nombre archivo']:<36} {f['Tipo contenido']:<20} {tipo_a:<18} {ok}  {f['ERP']}")

    xlsx_path = Path(f"contenidos_desde_{desde}.xlsx")
    print()
    guardar_excel(filas, xlsx_path)

    sin_arch = sum(1 for f in filas if not f["Tipo archivo"])
    mismatch = sum(1 for f in filas if f["Tipo archivo"] and
                   f["Tipo archivo"].lower() != f["Tipo contenido"].lower())

    print(f"\n{'=' * 62}")
    print(f"  Contenidos encontrados : {len(filas)}")
    print(f"  Sin archivo adjunto    : {sin_arch}")
    print(f"  Tipo no coincide       : {mismatch}")
    print(f"  Desde                  : {desde}")
    print(f"  Excel                  : {xlsx_path.resolve()}")
    print("=" * 62)


if __name__ == "__main__":
    main()
