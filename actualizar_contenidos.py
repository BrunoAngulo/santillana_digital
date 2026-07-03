#!/usr/bin/env python3
"""
Actualiza el nombre de contenidos en Compartir Conocimientos a partir de un
Excel editado (generado previamente por listar_contenidos.py).

Flujo:
  1. Lee el Excel: busca columnas "GUID" y "Nombre contenido"
  2. Por cada fila, GET el contenido actual desde la API
  3. Si el nombre en el Excel difiere del actual → PUT para actualizar
  4. Genera un log con los cambios realizados

Uso:
  python actualizar_contenidos.py contenidos_desde_2026-06-25.xlsx
  python actualizar_contenidos.py   (pide el archivo interactivamente)
"""

import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from tqdm import tqdm
except ImportError:
    print("Instalando dependencias...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "requests", "openpyxl", "tqdm"])
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from tqdm import tqdm


API_BASE    = "https://compartirconocimientos-pe.santillana.com"
MAX_WORKERS = 5   # PUT requests paralelos (conservador para no saturar la API)


# ── Sesión ────────────────────────────────────────────────────────────────────

def build_session(token: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "Authorization": token.strip(),
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Origin":   "https://publisher.compartirconocimientos-pe.santillana.com",
        "Referer":  "https://publisher.compartirconocimientos-pe.santillana.com/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/149.0.0.0 Safari/537.36 Edg/149.0.0.0"
        ),
    })
    return s


# ── Lectura del Excel editado ─────────────────────────────────────────────────

def leer_excel_cambios(xlsx_path: Path) -> list[dict]:
    """
    Lee el Excel y devuelve filas con al menos "guid" y "nombre_nuevo".
    Busca las columnas "GUID" y "Nombre contenido" por cabecera (no por posición),
    así funciona aunque el usuario haya reordenado columnas.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Leer cabeceras de la primera fila
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    try:
        col_guid   = headers.index("GUID")
        col_nombre = headers.index("Nombre contenido")
    except ValueError as e:
        sys.exit(f"\n[ERROR] Columna no encontrada en el Excel: {e}\n"
                 "Asegúrate de usar el Excel generado por listar_contenidos.py")

    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        guid   = str(row[col_guid]   or "").strip()
        nombre = str(row[col_nombre] or "").strip()
        if not guid or not nombre:
            continue
        filas.append({"guid": guid, "nombre_nuevo": nombre})

    return filas


# ── API: obtener y actualizar contenido ──────────────────────────────────────

def get_contenido(session: requests.Session, guid: str) -> dict | None:
    try:
        r = session.get(f"{API_BASE}/api/cms/contents/{guid}", timeout=30)
        r.raise_for_status()
        return r.json().get("data")
    except requests.RequestException as e:
        tqdm.write(f"  [ERROR GET] {guid}: {e}")
        return None


def actualizar_nombre(session: requests.Session, existente: dict,
                      nombre_nuevo: str) -> bool:
    """
    PUT /api/cms/contents/{guid} manteniendo todos los campos actuales,
    solo sobreescribiendo el campo "name".
    """
    guid = existente["guid"]

    payload = {
        "collections":          [c["guid"] for c in existente.get("collections", [])],
        "customTags":           [],
        "dependencies":         [],
        "description":          existente.get("description") or "",
        "didacticTypes":        [],
        "disciplines":          [d["discipline_guid"]      for d in existente.get("disciplines",     [])],
        "educationLevels":      [e["education_level_guid"] for e in existente.get("educationLevels",  [])],
        "educationYears":       [a["education_year_guid"]  for a in existente.get("educationYears",   [])],
        "encoded_transcription_url": existente.get("encoded_transcription_url"),
        "erp_id":               existente.get("erp_id", ""),
        "guid":                 guid,
        "is_available_offline": existente.get("is_available_offline", 1),
        "is_downloadable":      existente.get("is_downloadable", 0),
        "is_public":            existente.get("is_public", 0),
        "is_teacher_only":      existente.get("is_teacher_only", 0),
        "langs":                [l["id"] for l in existente.get("langs", [])],
        "learningObjectives":   [],
        "mobile_friendly":      existente.get("mobile_friendly", 1),
        "name":                 nombre_nuevo,      # ← único campo que cambia
        "publications":         [],
        "status":               existente.get("status", "active"),
        "tags":                 [],
        "topics":               [],
        "transcription_bundle": existente.get("transcription_bundle"),
        "transcription_url":    existente.get("transcription_url") or "",
        "type_guid":            existente.get("type_guid", ""),
    }

    try:
        r = session.put(f"{API_BASE}/api/cms/contents/{guid}", json=payload, timeout=30)
        r.raise_for_status()
        return r.json().get("status") == "success"
    except requests.RequestException as e:
        tqdm.write(f"  [ERROR PUT] {guid}: {e}")
        return False


# ── Worker por contenido ──────────────────────────────────────────────────────

def _procesar_fila(token: str, fila: dict) -> dict:
    """Crea su propia sesión (thread-safe). Devuelve dict con resultado."""
    session     = build_session(token)
    guid        = fila["guid"]
    nombre_nuevo = fila["nombre_nuevo"]
    resultado   = {
        "guid":          guid,
        "nombre_nuevo":  nombre_nuevo,
        "nombre_actual": "",
        "estado":        "ERROR",
        "detalle":       "",
    }

    existente = get_contenido(session, guid)
    if not existente:
        resultado["detalle"] = "No se pudo obtener el contenido"
        return resultado

    nombre_actual = existente.get("name", "")
    resultado["nombre_actual"] = nombre_actual

    if nombre_actual == nombre_nuevo:
        resultado["estado"]  = "SIN_CAMBIO"
        resultado["detalle"] = "Nombre idéntico"
        return resultado

    ok = actualizar_nombre(session, existente, nombre_nuevo)
    if ok:
        resultado["estado"]  = "ACTUALIZADO"
        resultado["detalle"] = f'"{nombre_actual}" → "{nombre_nuevo}"'
    else:
        resultado["detalle"] = "PUT falló"

    return resultado


# ── Log Excel ─────────────────────────────────────────────────────────────────

def guardar_log(resultados: list[dict], log_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cambios"

    headers  = ["GUID", "Nombre actual", "Nombre nuevo", "Estado", "Detalle"]
    fill_h   = PatternFill("solid", fgColor="1F4E79")
    font_h   = Font(bold=True, color="FFFFFF", size=11)
    fill_ok  = PatternFill("solid", fgColor="C6EFCE")
    fill_err = PatternFill("solid", fgColor="FFC7CE")
    fill_nc  = PatternFill("solid", fgColor="D9D9D9")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_h
        c.font = font_h
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(resultados, 2):
        ws.cell(row=i, column=1, value=r["guid"])
        ws.cell(row=i, column=2, value=r["nombre_actual"])
        ws.cell(row=i, column=3, value=r["nombre_nuevo"])
        ws.cell(row=i, column=4, value=r["estado"])
        ws.cell(row=i, column=5, value=r["detalle"])

        if r["estado"] == "ACTUALIZADO":
            fill = fill_ok
        elif r["estado"] == "ERROR":
            fill = fill_err
        else:
            fill = fill_nc
        for col in range(1, 6):
            ws.cell(row=i, column=col).fill = fill

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 60

    wb.save(log_path)
    print(f"  Log guardado: {log_path.resolve()}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  Actualizador de nombres - Santillana Digital")
    print("=" * 62)
    print()

    if len(sys.argv) > 1:
        xlsx_str = sys.argv[1].strip().strip('"').strip("'")
    else:
        xlsx_str = input("Ruta del Excel editado > ").strip().strip('"').strip("'")

    xlsx_path = Path(xlsx_str)
    if not xlsx_path.exists():
        sys.exit(f"[ERROR] El archivo no existe: {xlsx_path}")

    token = input("\nToken de autorización (Bearer ...) > ").strip()
    if not token:
        sys.exit("No se ingresó token.")

    print(f"\nLeyendo {xlsx_path.name}...")
    filas = leer_excel_cambios(xlsx_path)
    print(f"  {len(filas)} filas encontradas con GUID y nombre.")

    if not filas:
        sys.exit("No hay filas válidas para procesar.")

    print(f"\nProcesando (máximo {MAX_WORKERS} simultáneos)...\n")
    resultados = []
    barra = tqdm(total=len(filas), desc="Procesando", unit="item", ncols=80)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_procesar_fila, token, f): f for f in filas}
        for future in as_completed(futures):
            r = future.result()
            resultados.append(r)
            if r["estado"] == "ACTUALIZADO":
                tqdm.write(f"  [ACTUALIZADO] {r['detalle']}")
            elif r["estado"] == "SIN_CAMBIO":
                tqdm.write(f"  [=] {r['nombre_actual']}")
            else:
                tqdm.write(f"  [ERROR] {r['guid']}: {r['detalle']}")
            barra.update(1)

    barra.close()

    # Log
    from datetime import datetime
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = xlsx_path.parent / f"log_actualizacion_{ts}.xlsx"
    print()
    guardar_log(resultados, log_path)

    act = sum(1 for r in resultados if r["estado"] == "ACTUALIZADO")
    nc  = sum(1 for r in resultados if r["estado"] == "SIN_CAMBIO")
    err = sum(1 for r in resultados if r["estado"] == "ERROR")

    print(f"\n{'=' * 62}")
    print(f"  Actualizados  : {act}")
    print(f"  Sin cambio    : {nc}")
    print(f"  Errores       : {err}")
    print(f"  Log           : {log_path}")
    print("=" * 62)


if __name__ == "__main__":
    main()
