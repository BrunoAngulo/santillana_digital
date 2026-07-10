#!/usr/bin/env python3
"""
Sube archivos de SantillanaDigital_Docs/ al CMS y los vincula al LMS.

Estructura esperada:
  SantillanaDigital_Docs/
  ├── Ciencia y Tecnología 1.er grado - Secundaria/
  │   ├── Documentos curriculares/
  │   │   ├── Propósitos de aprendizaje U1.pdf
  │   │   └── ...
  │   └── Guía metodológica/
  │       └── ...
  └── ...

Flujo por producto:
  1. Por cada archivo: busca en CMS por nombre del stem exacto.
     - Si existe → usa su GUID.
     - Si no existe → sube al CMS con los metadatos del lote.
  2. Busca el curso LMS de forma interactiva.
  3. Por cada subcarpeta (módulo):
     - Crea un módulo LMS con el nombre de la carpeta (si no existe ya).
     - Crea una sección "Recursos" dentro del módulo.
     - Vincula los contenidos CMS a esa sección.
  4. Genera log Excel.

Uso:
  python subir_docs_santillana.py
  python subir_docs_santillana.py "C:\\...\\SantillanaDigital_Docs"
"""

import sys
import re
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from tqdm import tqdm
except ImportError:
    print("Instalando dependencias...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "requests", "openpyxl", "tqdm"])
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from tqdm import tqdm


# ── Constantes ───────────────────────────────────────────────────────────────

API_BASE = "https://compartirconocimientos-pe.santillana.com"

TYPE_GUID = {
    ".pdf":  "CTTY_05",
    ".zip":  "CTTY_08",
    ".docx": "CTTY_12",
    ".doc":  "CTTY_12",
    ".pptx": "CTTY_12",
    ".html": "CTTY_08",
    ".htm":  "CTTY_08",
}

EXTENSIONES_VALIDAS = set(TYPE_GUID.keys())

_RETRY_STATUS  = {502, 503, 504}
_RETRY_DELAYS  = [3, 6, 12]
MAX_WORKERS    = 5
_zip_sem       = threading.Semaphore(2)

_YEARS_FALLBACK = {
    "00000000-0000-1000-0000-000000000039": [
        {"guid": "00000000-0000-1000-0000-000000000119", "name": "1.er grado - Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000120", "name": "2.do grado - Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000121", "name": "3.er grado - Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000122", "name": "4.to grado - Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000123", "name": "5.to grado - Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000124", "name": "6.to grado - Primaria"},
    ],
    "00000000-0000-1000-0000-000000000040": [
        {"guid": "00000000-0000-1000-0000-000000000126", "name": "1.er año - Secundaria"},
        {"guid": "00000000-0000-1000-0000-000000000127", "name": "2.do año - Secundaria"},
        {"guid": "00000000-0000-1000-0000-000000000128", "name": "3.er año - Secundaria"},
        {"guid": "00000000-0000-1000-0000-000000000129", "name": "4.to año - Secundaria"},
        {"guid": "00000000-0000-1000-0000-000000000130", "name": "5.to año - Secundaria"},
    ],
}


# ── HTTP ─────────────────────────────────────────────────────────────────────

def build_session(token: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "Authorization": token.strip(),
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Origin":   "https://publisher.compartirconocimientos-pe.santillana.com",
        "Referer":  "https://publisher.compartirconocimientos-pe.santillana.com/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/149.0.0.0 Safari/537.36"
        ),
    })
    return s


def _raise_with_body(r):
    try:
        r.raise_for_status()
    except requests.HTTPError as e:
        raise requests.HTTPError(f"{e} — {r.text[:300]}", response=r) from None


def _request_with_retry(fn, *args, **kwargs):
    for intento, delay in enumerate([0] + _RETRY_DELAYS, 1):
        if delay:
            time.sleep(delay)
        try:
            r = fn(*args, **kwargs)
            if r.status_code in _RETRY_STATUS:
                tqdm.write(f"  [RETRY {intento}/4] {r.status_code} — reintentando...")
                continue
            _raise_with_body(r)
            return r
        except requests.ConnectionError:
            if intento <= len(_RETRY_DELAYS):
                tqdm.write(f"  [RETRY {intento}/4] ConnectionError — reintentando...")
                continue
            raise
    _raise_with_body(r)
    return r


def api_get(session, path: str, params=None):
    r = _request_with_retry(session.get, f"{API_BASE}{path}",
                             params=params, timeout=30)
    return r.json()


def api_post_json(session, path: str, payload: dict):
    r = _request_with_retry(session.post, f"{API_BASE}{path}",
                             json=payload, timeout=30)
    return r.json()


def api_post_files(session, endpoint: str, files: dict) -> dict:
    hdrs = {k: v for k, v in session.headers.items()
            if k.lower() != "content-type"}
    r = requests.post(endpoint if endpoint.startswith("http")
                      else f"{API_BASE}{endpoint}",
                      files=files, headers=hdrs, timeout=(30, 1800))
    _raise_with_body(r)
    return r.json()


def api_put(session, path: str, payload: dict):
    r = _request_with_retry(session.put, f"{API_BASE}{path}",
                             json=payload, timeout=30)
    return r.json()


# ── Menús interactivos ────────────────────────────────────────────────────────

def menu(titulo: str, opciones: list[dict],
         key_label: str = "name", key_value: str = "guid") -> str:
    print(f"\n{titulo}")
    print("─" * 50)
    for i, op in enumerate(opciones, 1):
        print(f"  {i:>2}. {op[key_label]}")
    while True:
        entrada = input("Elige número > ").strip()
        if entrada.isdigit() and 1 <= int(entrada) <= len(opciones):
            elegido = opciones[int(entrada) - 1]
            print(f"  ✓ {elegido[key_label]}")
            return elegido[key_value]
        print(f"  Número entre 1 y {len(opciones)}")


# ── Catálogos ─────────────────────────────────────────────────────────────────

def fetch_education_levels(session) -> list[dict]:
    try:
        data  = api_get(session, "/api/cms/education-levels")
        items = data.get("data", data)
        if isinstance(items, list):
            return [{"guid": x.get("guid", x.get("education_level_guid")),
                     "name": x.get("name", x.get("education_level_name", str(x)))}
                    for x in items]
    except Exception:
        pass
    return [
        {"guid": "00000000-0000-1000-0000-000000000038", "name": "Inicial"},
        {"guid": "00000000-0000-1000-0000-000000000039", "name": "Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000040", "name": "Secundaria"},
    ]


def fetch_education_years(session, level_guid: str) -> list[dict]:
    try:
        data  = api_get(session, "/api/cms/education-years",
                        params={"education_level_guid": level_guid})
        items = data.get("data", data)
        if isinstance(items, list) and items:
            return [{"guid": x.get("guid", x.get("education_year_guid")),
                     "name": x.get("name", x.get("education_year_name", str(x)))}
                    for x in items]
    except Exception:
        pass
    return _YEARS_FALLBACK.get(level_guid, [])


def fetch_disciplines(session) -> list[dict]:
    try:
        data  = api_get(session, "/api/cms/disciplines")
        items = data.get("data", data)
        if isinstance(items, list) and items:
            return [{"guid": x.get("guid", x.get("discipline_guid")),
                     "name": x.get("name", x.get("discipline_name", str(x)))}
                    for x in items]
    except Exception:
        pass
    return [
        {"guid": "00000000-0000-1000-0000-000000013146", "name": "Matemática"},
        {"guid": "00000000-0000-1000-0000-000000013147", "name": "Comunicación"},
        {"guid": "00000000-0000-1000-0000-000000013148", "name": "Personal Social"},
        {"guid": "00000000-0000-1000-0000-000000063304", "name": "Ciencia y Tecnología"},
    ]


def fetch_collections(session, search: str = "") -> list[dict]:
    data = api_get(session, "/api/cms/collections",
                   params={"pageSize": 100, "search": search})
    cols = data.get("data", {}).get("collections", [])
    return [{"guid": c["guid"], "name": c["collection"]} for c in cols]


# ── CMS: búsqueda y subida ────────────────────────────────────────────────────

def buscar_en_cms(session, stem: str) -> dict | None:
    """
    Busca en el CMS por stem del archivo.
    Coincidencia exacta en name o erp_id (case-insensitive).
    """
    data = api_get(session, "/api/cms/contents",
                   params={"search": stem, "pageSize": 20, "offset": 0, "page": 0,
                           "isEditorial": 1})
    contents = (data.get("data") or {}).get("contents", [])
    stem_l = stem.lower()
    for c in contents:
        if ((c.get("erp_id") or "").lower() == stem_l or
                (c.get("name") or "").lower() == stem_l):
            return c
    return None


def _mime(path: Path) -> str:
    return {
        ".pdf":  "application/pdf",
        ".zip":  "application/zip",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        ".html": "text/html",
    }.get(path.suffix.lower(), "application/octet-stream")


def subir_al_cms(session, filepath: Path,
                 level_guid: str, year_guid: str,
                 disc_guid: str, idioma: str,
                 col_guid: str) -> dict | None:
    """
    Crea contenido CMS, sube el archivo y aplica metadatos.
    Devuelve el dict completo del contenido (con guid) o None si falla.
    """
    erp_id    = filepath.stem
    nombre    = erp_id          # nombre visible = stem del archivo
    type_guid = TYPE_GUID.get(filepath.suffix.lower())
    if not type_guid:
        tqdm.write(f"  [SKIP] Extensión no soportada: {filepath.name}")
        return None

    is_zip = filepath.suffix.lower() == ".zip"
    if is_zip:
        _zip_sem.acquire()

    try:
        # 1. Crear entrada en CMS
        try:
            resp = api_post_json(session, "/api/cms/contents", {
                "description":          "",
                "erp_id":               erp_id,
                "is_available_offline": 1,
                "is_teacher_only":      1,
                "langs":                [],
                "name":                 nombre,
                "type_guid":            type_guid,
            })
            guid = resp["data"]["guid"]
        except requests.HTTPError as e:
            if "ER_DUP_ENTRY" not in str(e) and "1062" not in str(e):
                raise
            # ERP duplicado: buscar el existente
            existente = buscar_en_cms(session, erp_id)
            if not existente:
                raise RuntimeError(f"ERP duplicado pero no encontrado: {erp_id}")
            return existente

        # 2. Obtener endpoint de upload
        info   = api_get(session, f"/api/cms/contents/{guid}/content/upload")
        upload = info["data"]["data"]["upload"]
        token_up  = upload["token"]
        endpoint  = upload["endpoint"]

        # 3. Subir archivo
        for intento in range(3):
            try:
                with open(filepath, "rb") as f:
                    resp_up = api_post_files(session, endpoint, {
                        "file":  (filepath.name, f, _mime(filepath)),
                        "token": (None, token_up),
                    })
                if resp_up.get("status") != "success":
                    raise RuntimeError("Upload no devolvió success")
                break
            except (requests.RequestException, OSError) as e:
                if intento < 2:
                    time.sleep(10 * (2 ** intento))
                else:
                    raise

        # 4. Esperar procesamiento
        fin = time.time() + 600
        while time.time() < fin:
            estado = (api_get(session, f"/api/cms/contents/{guid}")
                      .get("data", {}).get("status", ""))
            if estado and estado != "processing":
                break
            time.sleep(5)

        # 5. Metadatos
        api_put(session, f"/api/cms/contents/{guid}", {
            "collections":          [col_guid],
            "customTags":           [],
            "dependencies":         [],
            "description":          "",
            "didacticTypes":        [],
            "disciplines":          [disc_guid],
            "educationLevels":      [level_guid],
            "educationYears":       [year_guid],
            "erp_id":               erp_id,
            "guid":                 guid,
            "is_available_offline": 1,
            "is_downloadable":      0,
            "is_public":            0,
            "is_teacher_only":      1,
            "langs":                [idioma],
            "learningObjectives":   [],
            "mobile_friendly":      1,
            "name":                 nombre,
            "status":               "active",
            "type_guid":            type_guid,
        })

        return {"guid": guid, "name": nombre, "erp_id": erp_id}

    finally:
        if is_zip:
            _zip_sem.release()


# ── LMS: cursos, módulos, secciones ──────────────────────────────────────────

def buscar_cursos(session, search: str) -> list[dict]:
    data = api_get(session, "/api/lms/courses", params={
        "offset": 0, "page": 0, "pageSize": 20,
        "isEditorial": 1, "search": search,
    })
    return data.get("data", {}).get("courses", [])


def get_course_items(session, course_guid: str) -> list[dict]:
    data = api_get(session, f"/api/front/courses/{course_guid}/items")
    return data.get("data", {}).get("items", [])


def html_a_texto(html: str) -> str:
    return re.sub(r"<[^>]+>", "", html or "").strip()


def crear_lesson(session, course_guid: str, nombre: str) -> str | None:
    html = f'<p><span style="font-size: 24px;">{nombre}</span></p>'
    data = api_post_json(session, f"/api/front/courses/{course_guid}/lessons", {
        "evaluation_period_id": "annual",
        "name":               html,
        "number_of_sessions": 1,
        "type":               "teacher",
    })
    d = data.get("data", {})
    return d.get("lesson_guid") or d.get("guid")


def crear_seccion(session, lesson_guid: str, nombre_seccion: str) -> str | None:
    r = session.post(f"{API_BASE}/api/front/lesson-items",
                     json={"lesson_guid": lesson_guid,
                           "section": nombre_seccion,
                           "section_type_guid": "default"},
                     timeout=30)
    try:
        _raise_with_body(r)
    except requests.HTTPError as e:
        tqdm.write(f"      [ERROR] crear_seccion: {e}")
        return None
    data = r.json()
    guid = (data.get("data") or {}).get("guid")
    if not guid:
        tqdm.write(f"      [WARN] crear_seccion sin guid: {data!s:.200}")
    return guid


def agregar_contenido(session, lesson_guid: str, parent_guid: str,
                      content: dict, nombre_visible: str) -> str | None:
    data = api_post_json(session, "/api/front/lesson-items", {
        "content_guid":         content["guid"],
        "description":          "",
        "disciplines":          [d["discipline_guid"]      for d in content.get("disciplines",    [])],
        "educationLevels":      [e["education_level_guid"] for e in content.get("educationLevels", [])],
        "educationYears":       [a["education_year_guid"]  for a in content.get("educationYears",  [])],
        "include_in_gradebook": 0,
        "is_embed":             1,
        "item_for":             "all",
        "learningObjectives":   [],
        "lesson_guid":          lesson_guid,
        "name":                 nombre_visible,
        "parent_guid":          parent_guid,
        "status":               "published",
        "teacher_notes":        None,
    })
    return (data.get("data") or {}).get("guid") or None


# ── Selección interactiva de curso ────────────────────────────────────────────

def seleccionar_curso(session, producto_nombre: str) -> dict | None:
    while True:
        term = input(f"\n  Búsqueda del curso LMS para '{producto_nombre}' > ").strip()
        if not term:
            print("  (saltando este producto)")
            return None
        cursos = buscar_cursos(session, term)
        if not cursos:
            print("  Sin resultados, intenta con otro término.")
            continue
        if len(cursos) == 1:
            c = cursos[0]
            print(f"  Curso: {c['name']}  [{c['guid']}]")
            return c
        print()
        for i, c in enumerate(cursos, 1):
            año  = c.get("education_year_name", "?")
            disc = c.get("discipline_name",     "?")
            print(f"    {i}. {c['name']} — {año} — {disc}")
        while True:
            r = input(f"  Selección (1-{len(cursos)}) o 0 para volver a buscar > ").strip()
            if r == "0":
                break
            if r.isdigit() and 1 <= int(r) <= len(cursos):
                return cursos[int(r) - 1]


# ── Escaneo de carpetas ───────────────────────────────────────────────────────

def escanear_producto(producto_path: Path) -> dict[str, list[Path]]:
    """
    Devuelve {nombre_carpeta → [archivos]} para un producto.
    Ignora subcarpetas vacías o sin archivos de tipo válido.
    """
    carpetas: dict[str, list[Path]] = {}
    for sub in sorted(producto_path.iterdir()):
        if not sub.is_dir():
            continue
        archivos = sorted(
            f for f in sub.iterdir()
            if f.is_file() and f.suffix.lower() in EXTENSIONES_VALIDAS
        )
        if archivos:
            carpetas[sub.name] = archivos
    return carpetas


# ── Log Excel ─────────────────────────────────────────────────────────────────

_FILL_HDR  = PatternFill("solid", fgColor="1F4E79")
_FONT_HDR  = Font(bold=True, color="FFFFFF", size=11)
_FILL_OK   = PatternFill("solid", fgColor="C6EFCE")
_FILL_UPLD = PatternFill("solid", fgColor="FFEB9C")
_FILL_ERR  = PatternFill("solid", fgColor="FFC7CE")
_FILL_SKIP = PatternFill("solid", fgColor="D9D9D9")


def guardar_log(entradas: list[dict], log_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Log"

    headers = ["Producto", "Sección", "Archivo", "Estado CMS",
               "GUID CMS", "Estado LMS", "GUID item LMS", "Detalle"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _FILL_HDR
        c.font = _FONT_HDR
        c.alignment = Alignment(horizontal="center")

    for i, e in enumerate(entradas, 2):
        ws.cell(row=i, column=1, value=e.get("producto",      ""))
        ws.cell(row=i, column=2, value=e.get("seccion",        ""))
        ws.cell(row=i, column=3, value=e.get("archivo",       ""))
        ws.cell(row=i, column=4, value=e.get("estado_cms",    ""))
        ws.cell(row=i, column=5, value=e.get("guid_cms",      ""))
        ws.cell(row=i, column=6, value=e.get("estado_lms",    ""))
        ws.cell(row=i, column=7, value=e.get("guid_item_lms", ""))
        ws.cell(row=i, column=8, value=e.get("detalle",       ""))

        estado = e.get("estado_cms", "")
        if estado == "SUBIDO":
            fill = _FILL_UPLD
        elif estado == "ENCONTRADO":
            fill = _FILL_OK
        elif estado in ("SKIP", "SIN_CURSO"):
            fill = _FILL_SKIP
        else:
            fill = _FILL_ERR

        for col in range(1, 9):
            ws.cell(row=i, column=col).fill = fill

    anchos = {"A": 40, "B": 30, "C": 45, "D": 12,
              "E": 38, "F": 12, "G": 38, "H": 45}
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    wb.save(log_path)
    print(f"  Log guardado: {log_path.resolve()}")


# ── Procesamiento de un producto ──────────────────────────────────────────────

def procesar_producto(session, token: str, producto_path: Path,
                      level_guid: str, year_guid: str,
                      disc_guid: str, idioma: str, col_guid: str) -> list[dict]:
    """
    Estructura LMS resultante:
      Módulo "Documentos"
        └── Sección "Documentos curriculares"  ← subcarpeta del producto
              └── archivo1, archivo2 ...
        └── Sección "Guía metodológica"
              └── ...
    """
    entradas: list[dict] = []
    nombre_producto = producto_path.name

    carpetas = escanear_producto(producto_path)
    if not carpetas:
        print("  [AVISO] Sin carpetas con archivos válidos.")
        return entradas

    total_archivos = sum(len(v) for v in carpetas.values())
    print(f"  Secciones: {list(carpetas.keys())}")
    print(f"  Total archivos: {total_archivos}")

    # ── Seleccionar curso LMS ─────────────────────────────────────────
    curso = seleccionar_curso(session, nombre_producto)
    if not curso:
        for sec, files in carpetas.items():
            for fp in files:
                entradas.append({
                    "producto":      nombre_producto,
                    "seccion":       sec,
                    "archivo":       fp.name,
                    "estado_cms":    "SIN_CURSO",
                    "guid_cms":      "",
                    "estado_lms":    "SIN_CURSO",
                    "guid_item_lms": "",
                    "detalle":       "Producto saltado sin seleccionar curso",
                })
        return entradas

    course_guid = curso["guid"]

    # ── Obtener o crear el módulo "Documentos" ────────────────────────
    items_existentes   = get_course_items(session, course_guid)
    modulos_existentes = {html_a_texto(it.get("lesson_name", "")).lower(): it
                          for it in items_existentes}

    if "documentos" in modulos_existentes:
        lesson_guid = modulos_existentes["documentos"].get("lesson_guid", "")
        tqdm.write(f"  [OK] Módulo 'Documentos' encontrado  [{lesson_guid}]")
    else:
        lesson_guid = crear_lesson(session, course_guid, "Documentos")
        if lesson_guid:
            tqdm.write(f"  [+] Módulo 'Documentos' creado  [{lesson_guid}]")
        else:
            print("  [ERROR] No se pudo encontrar ni crear el módulo 'Documentos'.")
            for sec, files in carpetas.items():
                for fp in files:
                    entradas.append({
                        "producto":      nombre_producto,
                        "seccion":       sec,
                        "archivo":       fp.name,
                        "estado_cms":    "ERROR",
                        "guid_cms":      "",
                        "estado_lms":    "ERROR",
                        "guid_item_lms": "",
                        "detalle":       "No se encontró ni creó módulo 'Documentos'",
                    })
            return entradas

    # ── Fase 1: CMS — buscar / subir en paralelo ──────────────────────
    print(f"\n  Fase 1: CMS — buscando / subiendo {total_archivos} archivos...")

    items_planos: list[dict] = []
    for sec, files in carpetas.items():
        for fp in files:
            items_planos.append({"seccion": sec, "filepath": fp})

    cms_results: dict[Path, dict] = {}

    def _procesar_archivo(item: dict) -> tuple[Path, str, dict | None, str]:
        fp = item["filepath"]
        try:
            existente = buscar_en_cms(session, fp.stem)
            if existente:
                return fp, "ENCONTRADO", existente, ""
            content = subir_al_cms(session, fp,
                                   level_guid, year_guid,
                                   disc_guid, idioma, col_guid)
            if content:
                return fp, "SUBIDO", content, ""
            return fp, "ERROR", None, "subir_al_cms devolvió None"
        except Exception as e:
            return fp, "ERROR", None, str(e)

    barra = tqdm(total=len(items_planos), desc="  CMS", unit="arch", ncols=80)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = {ex.submit(_procesar_archivo, it): it for it in items_planos}
        for fut in as_completed(futs):
            fp, estado, content, detalle = fut.result()
            cms_results[fp] = {"estado": estado, "content": content, "detalle": detalle}
            tqdm.write(f"    [{estado}] {fp.name}")
            barra.update(1)
    barra.close()

    # ── Fase 2: LMS — crear secciones dentro de "Documentos" ──────────
    print(f"\n  Fase 2: LMS — creando secciones en módulo 'Documentos'...")

    for sec, files in carpetas.items():
        # Una sección por subcarpeta dentro del módulo Documentos
        parent_guid = crear_seccion(session, lesson_guid, sec)
        if not parent_guid:
            tqdm.write(f"    [ERROR] No se pudo crear sección '{sec}'")
            for fp in files:
                r = cms_results.get(fp, {})
                entradas.append({
                    "producto":      nombre_producto,
                    "seccion":       sec,
                    "archivo":       fp.name,
                    "estado_cms":    r.get("estado", "ERROR"),
                    "guid_cms":      (r.get("content") or {}).get("guid", ""),
                    "estado_lms":    "ERROR",
                    "guid_item_lms": "",
                    "detalle":       f"No se pudo crear sección '{sec}'",
                })
            continue

        tqdm.write(f"    [+] Sección: {sec}  [{parent_guid}]")

        for fp in files:
            r          = cms_results.get(fp, {})
            estado_cms = r.get("estado", "ERROR")
            content    = r.get("content")
            detalle    = r.get("detalle", "")

            if not content or not content.get("guid"):
                tqdm.write(f"      [SKIP] {fp.name} — sin GUID CMS")
                entradas.append({
                    "producto":      nombre_producto,
                    "seccion":       sec,
                    "archivo":       fp.name,
                    "estado_cms":    estado_cms,
                    "guid_cms":      "",
                    "estado_lms":    "SKIP",
                    "guid_item_lms": "",
                    "detalle":       detalle or "Sin GUID CMS",
                })
                continue

            item_guid = agregar_contenido(
                session, lesson_guid, parent_guid,
                content, fp.stem,
            )
            if item_guid:
                tqdm.write(f"      [+] {fp.name}")
                estado_lms = "VINCULADO"
            else:
                tqdm.write(f"      [ERROR] {fp.name}")
                estado_lms = "ERROR"
                detalle    = "agregar_contenido falló"

            entradas.append({
                "producto":      nombre_producto,
                "seccion":       sec,
                "archivo":       fp.name,
                "estado_cms":    estado_cms,
                "guid_cms":      content.get("guid", ""),
                "estado_lms":    estado_lms,
                "guid_item_lms": item_guid or "",
                "detalle":       detalle,
            })
            time.sleep(0.1)

    return entradas


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 64)
    print("  Subir Docs Santillana → CMS + LMS")
    print("=" * 64)

    # Ruta base
    if len(sys.argv) > 1:
        base_str = sys.argv[1].strip().strip('"').strip("'")
    else:
        base_str = input(
            "\nRuta de SantillanaDigital_Docs/ > "
        ).strip().strip('"').strip("'")

    base_path = Path(base_str)
    if not base_path.is_dir():
        sys.exit(f"\n[ERROR] La carpeta no existe: {base_path}")

    # Token
    token   = input("\nToken de autorización (Bearer ...) > ").strip()
    session = build_session(token)

    # Detectar productos
    productos = sorted(p for p in base_path.iterdir() if p.is_dir())
    if not productos:
        sys.exit("No se encontraron subcarpetas en la ruta indicada.")

    print(f"\nProductos encontrados: {len(productos)}")
    for i, p in enumerate(productos, 1):
        print(f"  {i:>3}. {p.name}")

    selec = input("\n¿Procesar todos? [s] o números separados por comas > ").strip().lower()
    if selec not in ("", "s"):
        nums     = [int(x.strip()) - 1 for x in selec.split(",") if x.strip().isdigit()]
        productos = [productos[i] for i in nums if 0 <= i < len(productos)]

    # ── Metadatos CMS comunes o por producto ──────────────────────────
    print("\n" + "=" * 64)
    print("  ¿Usar los mismos metadatos CMS para todos los productos?")
    print("  [s] Sí — pedir una sola vez")
    print("  [n] No — pedir por cada producto")
    mismos = input("  Opción [s/n] > ").strip().lower()

    metadata_comun = None
    if mismos in ("", "s"):
        print("\n  Metadatos CMS comunes a todos los productos:")
        levels     = fetch_education_levels(session)
        level_guid = menu("Etapa educativa", levels)
        years      = fetch_education_years(session, level_guid)
        year_guid  = menu("Año / Serie", years) if years else input("GUID año > ").strip()
        discs      = fetch_disciplines(session)
        disc_guid  = menu("Asignatura / Disciplina", discs) if discs else input("GUID disc > ").strip()
        idioma     = "es"
        busq_col   = input("\nBúsqueda colección > ").strip()
        cols       = fetch_collections(session, busq_col)
        col_guid   = menu("Colección", cols) if cols else input("GUID colección > ").strip()
        metadata_comun = (level_guid, year_guid, disc_guid, idioma, col_guid)

    # ── Procesar cada producto ─────────────────────────────────────────
    todas_entradas: list[dict] = []

    for producto_path in productos:
        print(f"\n{'─' * 64}")
        print(f"  Producto: {producto_path.name}")

        if metadata_comun:
            level_guid, year_guid, disc_guid, idioma, col_guid = metadata_comun
        else:
            print("  Metadatos CMS para este producto:")
            levels     = fetch_education_levels(session)
            level_guid = menu("Etapa educativa", levels)
            years      = fetch_education_years(session, level_guid)
            year_guid  = menu("Año / Serie", years) if years else input("GUID año > ").strip()
            discs      = fetch_disciplines(session)
            disc_guid  = menu("Asignatura / Disciplina", discs) if discs else input("GUID disc > ").strip()
            idioma     = "es"
            busq_col   = input("\nBúsqueda colección > ").strip()
            cols       = fetch_collections(session, busq_col)
            col_guid   = menu("Colección", cols) if cols else input("GUID colección > ").strip()

        entradas = procesar_producto(
            session, token, producto_path,
            level_guid, year_guid, disc_guid, idioma, col_guid,
        )
        todas_entradas.extend(entradas)

    # ── Log ───────────────────────────────────────────────────────────
    if todas_entradas:
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = base_path / f"log_docs_{ts}.xlsx"
        print()
        guardar_log(todas_entradas, log_path)

    # ── Resumen ───────────────────────────────────────────────────────
    subidos    = sum(1 for e in todas_entradas if e["estado_cms"]  == "SUBIDO")
    encontr    = sum(1 for e in todas_entradas if e["estado_cms"]  == "ENCONTRADO")
    vinculados = sum(1 for e in todas_entradas if e["estado_lms"]  == "VINCULADO")
    errores    = sum(1 for e in todas_entradas if e["estado_cms"]  == "ERROR"
                     or e["estado_lms"] == "ERROR")

    print(f"\n{'=' * 64}")
    print("  RESUMEN FINAL")
    print(f"{'=' * 64}")
    print(f"  Archivos encontrados en CMS : {encontr}")
    print(f"  Archivos subidos al CMS     : {subidos}")
    print(f"  Contenidos vinculados al LMS: {vinculados}")
    print(f"  Errores                     : {errores}")
    print("=" * 64)


if __name__ == "__main__":
    main()
