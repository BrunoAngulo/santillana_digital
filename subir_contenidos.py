#!/usr/bin/env python3
"""
Subidor automático de contenidos a Compartir Conocimientos (Santillana).
Lee un Excel con la lista de archivos, los busca en la carpeta Completo/,
y los sube aplicando los mismos metadatos a todo el lote.
"""

import sys
import time
import json
import re
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime

try:
    import requests
    from tqdm import tqdm
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Instalando dependencias...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "requests", "tqdm", "openpyxl"])
    import requests
    from tqdm import tqdm
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment


# ── Constantes ──────────────────────────────────────────────────────────────

API_BASE = "https://compartirconocimientos-pe.santillana.com"

# Mapeo extensión → type_guid
# CTTY_05=PDF  CTTY_08=HTML Interactivo  CTTY_12=Office
TYPE_GUID = {
    ".pdf":  "CTTY_05",
    ".zip":  "CTTY_08",   # ZIP = HTML Interactivo (paquete HTML)
    ".docx": "CTTY_12",
    ".doc":  "CTTY_12",
    ".pptx": "CTTY_12",
    ".xlsx": "CTTY_12",
    ".html": "CTTY_08",
    ".htm":  "CTTY_08",
}

# is_teacher_only según "disponible para"
DISPONIBLE_MAP = {
    "1": (0, "Docentes y Estudiantes"),
    "2": (1, "Solo Docentes"),
}

IDIOMAS_MAP = {
    "1": "es",
    "2": "en",
    "3": "pt",
}


# ── Sesión HTTP ─────────────────────────────────────────────────────────────

def build_session(token: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "Authorization": token.strip(),
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://publisher.compartirconocimientos-pe.santillana.com",
        "Referer": "https://publisher.compartirconocimientos-pe.santillana.com/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/149.0.0.0 Safari/537.36 Edg/149.0.0.0"
        ),
    })
    return s


def _raise_with_body(r):
    try:
        r.raise_for_status()
    except requests.HTTPError as e:
        raise requests.HTTPError(f"{e} — {r.text[:300]}", response=r) from None


def api_get(session, path: str, params=None):
    url = f"{API_BASE}{path}"
    r = session.get(url, params=params, timeout=30)
    _raise_with_body(r)
    return r.json()


def api_post(session, path: str, json_data=None, files=None, headers_extra=None):
    url = f"{API_BASE}{path}"
    if files:
        # Multipart upload: quitar Content-Type para que requests lo ponga con boundary
        hdrs = {k: v for k, v in session.headers.items() if k.lower() != "content-type"}
        if headers_extra:
            hdrs.update(headers_extra)
        r = requests.post(url, files=files, headers=hdrs, timeout=(30, 1800))
    else:
        r = session.post(url, json=json_data, timeout=30)
    _raise_with_body(r)
    return r.json()


def api_put(session, path: str, json_data: dict):
    url = f"{API_BASE}{path}"
    r = session.put(url, json=json_data, timeout=30)
    _raise_with_body(r)
    return r.json()


# ── Menús interactivos ──────────────────────────────────────────────────────

def menu(titulo: str, opciones: list[dict], key_label="name", key_value="guid") -> str:
    """
    Muestra un menú numerado y devuelve el valor del campo key_value elegido.
    opciones: lista de dicts con al menos key_label y key_value.
    """
    print(f"\n{titulo}")
    print("─" * 50)
    for i, op in enumerate(opciones, 1):
        print(f"  {i:>2}. {op[key_label]}")
    print()
    while True:
        entrada = input("Elige número > ").strip()
        if entrada.isdigit() and 1 <= int(entrada) <= len(opciones):
            elegido = opciones[int(entrada) - 1]
            print(f"  ✓ Seleccionado: {elegido[key_label]}")
            return elegido[key_value]
        print(f"  Ingresa un número entre 1 y {len(opciones)}")


def menu_simple(titulo: str, opciones: dict) -> tuple:
    """opciones: {clave: (valor, etiqueta)}. Devuelve (valor, etiqueta)."""
    print(f"\n{titulo}")
    print("─" * 50)
    for k, (_, etiqueta) in opciones.items():
        print(f"  {k}. {etiqueta}")
    print()
    while True:
        entrada = input("Elige número > ").strip()
        if entrada in opciones:
            valor, etiqueta = opciones[entrada]
            print(f"  ✓ Seleccionado: {etiqueta}")
            return valor, etiqueta
        print(f"  Opción no válida.")


# ── Fetch de catálogos desde la API ─────────────────────────────────────────

def fetch_education_levels(session) -> list[dict]:
    try:
        data = api_get(session, "/api/cms/education-levels")
        items = data.get("data", data)
        if isinstance(items, list):
            return [{"guid": x.get("guid", x.get("education_level_guid")),
                     "name": x.get("name", x.get("education_level_name", str(x)))}
                    for x in items]
    except Exception:
        pass
    # Fallback con los valores conocidos
    return [
        {"guid": "00000000-0000-1000-0000-000000000038", "name": "Inicial"},
        {"guid": "00000000-0000-1000-0000-000000000039", "name": "Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000040", "name": "Secundaria"},
    ]


_YEARS_FALLBACK = {
    # Primaria (nivel 039)
    "00000000-0000-1000-0000-000000000039": [
        {"guid": "00000000-0000-1000-0000-000000000119", "name": "1.er grado - Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000120", "name": "2.do grado - Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000121", "name": "3.er grado - Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000122", "name": "4.to grado - Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000123", "name": "5.to grado - Primaria"},
        {"guid": "00000000-0000-1000-0000-000000000124", "name": "6.to grado - Primaria"},
    ],
    # Secundaria (nivel 040)
    "00000000-0000-1000-0000-000000000040": [
        {"guid": "00000000-0000-1000-0000-000000000126", "name": "1.er año - Secundaria"},
        {"guid": "00000000-0000-1000-0000-000000000127", "name": "2.do año - Secundaria"},
        {"guid": "00000000-0000-1000-0000-000000000128", "name": "3.er año - Secundaria"},
        {"guid": "00000000-0000-1000-0000-000000000129", "name": "4.to año - Secundaria"},
        {"guid": "00000000-0000-1000-0000-000000000130", "name": "5.to año - Secundaria"},
    ],
}


def fetch_education_years(session, level_guid: str) -> list[dict]:
    try:
        data = api_get(session, "/api/cms/education-years",
                       params={"education_level_guid": level_guid})
        items = data.get("data", data)
        if isinstance(items, list) and items:
            return [{"guid": x.get("guid", x.get("education_year_guid")),
                     "name": x.get("name", x.get("education_year_name", str(x)))}
                    for x in items]
    except Exception:
        pass
    # Fallback con GUIDs conocidos
    return _YEARS_FALLBACK.get(level_guid, [])


_DISCIPLINES_FALLBACK = [
    {"guid": "00000000-0000-1000-0000-000000013146", "name": "Matemática"},
    {"guid": "00000000-0000-1000-0000-000000013147", "name": "Comunicación"},
    {"guid": "00000000-0000-1000-0000-000000013148", "name": "Personal Social"},
    {"guid": "00000000-0000-1000-0000-000000063304", "name": "Ciencia y Tecnología"},
]


def fetch_disciplines(session) -> list[dict]:
    try:
        data = api_get(session, "/api/cms/disciplines")
        items = data.get("data", data)
        if isinstance(items, list) and items:
            return [{"guid": x.get("guid", x.get("discipline_guid")),
                     "name": x.get("name", x.get("discipline_name", str(x)))}
                    for x in items]
    except Exception:
        pass
    return _DISCIPLINES_FALLBACK


def fetch_collections(session, search: str = "") -> list[dict]:
    data = api_get(session, "/api/cms/collections",
                   params={"pageSize": 100, "search": search})
    cols = data.get("data", {}).get("collections", [])
    return [{"guid": c["guid"], "name": c["collection"]} for c in cols]


# ── Búsqueda de contenido existente ─────────────────────────────────────────

def buscar_contenido_por_erp(session, erp_id: str) -> dict | None:
    """GET /api/cms/contents?search={erp_id} → devuelve el dict del contenido si existe."""
    data = api_get(session, "/api/cms/contents",
                   params={"search": erp_id, "pageSize": 10, "offset": 0, "page": 0})
    for item in (data.get("data") or {}).get("contents", []):
        if item.get("erp_id") == erp_id:
            return item
    return None


# ── Lógica de subida ─────────────────────────────────────────────────────────

def crear_contenido(session, nombre_visible: str, erp_id: str, type_guid: str,
                    is_teacher_only: int) -> str:
    """POST /api/cms/contents → devuelve el guid del contenido creado."""
    payload = {
        "description": "",
        "erp_id": erp_id,
        "is_available_offline": 1,
        "is_teacher_only": is_teacher_only,
        "langs": [],
        "name": nombre_visible,
        "type_guid": type_guid,
    }
    resp = api_post(session, "/api/cms/contents", json_data=payload)
    return resp["data"]["guid"]


def obtener_upload_info(session, guid: str) -> dict:
    """GET /api/cms/contents/{guid}/content/upload → token y endpoint."""
    resp = api_get(session, f"/api/cms/contents/{guid}/content/upload")
    upload = resp["data"]["data"]["upload"]
    return {"token": upload["token"], "endpoint": upload["endpoint"]}


def esperar_procesado(session, guid: str, timeout_seg: int = 600) -> bool:
    """Polling a GET /api/cms/contents/{guid} hasta que el estado no sea 'processing'."""
    fin = time.time() + timeout_seg
    while time.time() < fin:
        try:
            data = api_get(session, f"/api/cms/contents/{guid}")
            estado = (data.get("data") or {}).get("status", "")
            if estado and estado != "processing":
                return True
        except requests.RequestException:
            pass
        time.sleep(5)
    return False


def subir_archivo(session, endpoint: str, upload_token: str, filepath: Path) -> bool:
    """POST multipart al endpoint de files-storage, con reintentos."""
    mime = _mime(filepath)
    for intento in range(3):
        try:
            with open(filepath, "rb") as f:
                resp = api_post(session, endpoint, files={
                    "file": (filepath.name, f, mime),
                    "token": (None, upload_token),
                })
            return resp.get("status") == "success"
        except (requests.RequestException, OSError) as e:
            if intento < 2:
                espera = 10 * (2 ** intento)  # 10 s, 20 s
                tqdm.write(f"  [reintento {intento + 2}/3 en {espera}s] {filepath.name}: {e}")
                time.sleep(espera)
            else:
                raise


def _mime(path: Path) -> str:
    ext = path.suffix.lower()
    return {
        ".pdf":  "application/pdf",
        ".zip":  "application/zip",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".html": "text/html",
    }.get(ext, "application/octet-stream")


def resetear_archivo_contenido(session, guid: str) -> bool:
    """PUT /api/cms/contents/{guid} vaciando url/bundle para preparar re-upload."""
    payload = {
        "guid":                      guid,
        "bundle":                    "",
        "url":                       "",
        "url_public_viewer":         "",
        "storage_status":            None,
        "transcription_url":         None,
        "encoded_transcription_url": None,
        "transcription_bundle":      None,
        "subtitles":                 [],
        "audiodescriptions":         [],
    }
    resp = api_put(session, f"/api/cms/contents/{guid}", json_data=payload)
    return resp.get("status") == "success"


def actualizar_metadata(session, guid: str, nombre_visible: str, erp_id: str,
                        type_guid: str, is_teacher_only: int,
                        education_levels: list, education_years: list,
                        disciplines: list, langs: list,
                        collections: list) -> bool:
    payload = {
        "collections":          collections,
        "customTags":           [],
        "dependencies":         [],
        "description":          "",
        "didacticTypes":        [],
        "disciplines":          disciplines,
        "educationLevels":      education_levels,
        "educationYears":       education_years,
        "encoded_transcription_url": None,
        "erp_id":               erp_id,
        "guid":                 guid,
        "is_available_offline": 1,
        "is_downloadable":      0,
        "is_public":            0,
        "is_teacher_only":      is_teacher_only,
        "langs":                langs,
        "learningObjectives":   [],
        "mobile_friendly":      1,
        "name":                 nombre_visible,
        "publications":         [],
        "status":               "active",
        "tags":                 [],
        "topics":               [],
        "transcription_bundle": None,
        "transcription_url":    "",
        "type_guid":            type_guid,
    }
    resp = api_put(session, f"/api/cms/contents/{guid}", json_data=payload)
    return resp.get("status") == "success"


# ── Lectura del Excel ────────────────────────────────────────────────────────

_PARA_MAP = {
    "syd": 0,   # SyD → Docentes y Estudiantes
    "d":   1,   # D   → Solo Docentes
}


def leer_excel(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    archivos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        # Rellenar hasta 7 columnas para no fallar si G no existe
        cols           = list(row) + [None] * 7
        nombre_visible = cols[0]
        nombre_archivo = cols[1]
        ruta_dest      = cols[2]
        tipo           = cols[3]
        carpeta        = cols[4]
        carpeta_fuente = cols[5]
        para_col       = cols[6]   # columna G: "SyD" | "D" | vacío

        if not nombre_archivo:
            continue
        ext       = Path(str(nombre_archivo)).suffix.lower()
        erp_id    = Path(str(nombre_archivo)).stem
        type_guid = TYPE_GUID.get(ext)
        if not type_guid:
            tqdm.write(f"  [AVISO] Extensión desconocida ignorada: {nombre_archivo}")
            continue

        # Mapear columna G → is_teacher_only (None si no hay valor)
        para_str         = str(para_col).strip().lower() if para_col else ""
        is_teacher_excel = _PARA_MAP.get(para_str)   # 0, 1 ó None

        archivos.append({
            "nombre_visible":   str(nombre_visible).strip().rstrip('.'),
            "nombre_archivo":   str(nombre_archivo).strip(),
            "erp_id":           erp_id,
            "type_guid":        type_guid,
            "carpeta":          str(carpeta).strip() if carpeta else "",
            "tipo":             str(tipo).strip() if tipo else "ARCHIVO",
            "carpeta_fuente":   str(carpeta_fuente).strip() if carpeta_fuente else "sin_categoria",
            "is_teacher_excel": is_teacher_excel,   # None si columna G vacía/ausente
        })
    return archivos


# ── Log de resultados ────────────────────────────────────────────────────────

def guardar_log(resultados: list[dict], ruta_log: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    headers = ["Archivo", "Nombre visible", "GUID", "Estado", "Error", "URL viewer"]
    fill_h = PatternFill("solid", fgColor="1F4E79")
    font_h = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_h
        c.font = font_h
        c.alignment = Alignment(horizontal="center")

    fill_ok   = PatternFill("solid", fgColor="C6EFCE")
    fill_upd  = PatternFill("solid", fgColor="FFEB9C")
    fill_err  = PatternFill("solid", fgColor="FFC7CE")
    for i, r in enumerate(resultados, 2):
        if r["estado"] == "OK":
            fill = fill_ok
        elif r["estado"] == "ACTUALIZADO":
            fill = fill_upd
        else:
            fill = fill_err
        valores = [r["archivo"], r["nombre_visible"], r.get("guid", ""),
                   r["estado"], r.get("error", ""), r.get("url_viewer", "")]
        for col, val in enumerate(valores, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 70
    wb.save(ruta_log)


MAX_WORKERS  = 10  # máximo de uploads simultáneos en total
MAX_ZIP_SEM  = 3   # de esos 10, cuántos ZIPs pueden subir a la vez
_zip_sem     = threading.Semaphore(MAX_ZIP_SEM)


def _subir_item(token: str, item: dict, level_guid: str, year_guid: str,
                disc_guid: str, idioma_val: str, col_guid: str) -> dict:
    """Sube un archivo completo (crear → upload → metadata). Crea su propia sesión.
    Los ZIPs adquieren _zip_sem para no saturar el ancho de banda."""
    is_zip = item["ruta"].suffix.lower() == ".zip"
    if is_zip:
        _zip_sem.acquire()
    session = build_session(token)
    nombre    = item["nombre_visible"]
    erp_id    = item["erp_id"]
    type_g    = item["type_guid"]
    filepath  = item["ruta"]
    resultado = {"archivo": filepath.name, "nombre_visible": nombre,
                 "estado": "ERROR", "guid": "", "error": "", "url_viewer": ""}
    try:
        # ── Intentar crear contenido ─────────────────────────────────
        guid         = None
        es_duplicado = False
        try:
            guid = crear_contenido(session, nombre, erp_id, type_g, item["is_teacher_only"])
        except requests.HTTPError as e:
            if "ER_DUP_ENTRY" not in str(e) and "1062" not in str(e):
                raise
            es_duplicado = True
            tqdm.write(f"  [ACTUALIZANDO] {filepath.name} — ya existe, mergeando metadata...")

        viewer_url = lambda g: (
            f"https://publisher.compartirconocimientos-pe.santillana.com/content/{g}/1"
        )

        if es_duplicado:
            # ── Contenido ya existe: resetear archivo y re-subir ─────
            existente = buscar_contenido_por_erp(session, erp_id)
            if not existente:
                raise RuntimeError(f"ER_DUP_ENTRY pero no se encontró erp_id={erp_id}")
            guid = existente["guid"]
            resultado["guid"] = guid

            # Extraer GUIDs existentes (objetos anidados de la API)
            lvls_exist = {x["education_level_guid"] for x in existente.get("educationLevels", [])}
            yrs_exist  = {x["education_year_guid"]  for x in existente.get("educationYears",  [])}
            disc_exist = {x["discipline_guid"]      for x in existente.get("disciplines",     [])}
            cols_exist = {x["guid"]                 for x in existente.get("collections",     [])}
            lang_exist = {x["id"]                   for x in existente.get("langs",           [])}

            lvls      = list(lvls_exist | {level_guid})
            yrs       = list(yrs_exist  | {year_guid})
            disc      = list(disc_exist | {disc_guid})
            cols      = list(cols_exist | {col_guid})
            langs_act = list(lang_exist | {idioma_val})

            # 1. Vaciar el archivo anterior
            tqdm.write(f"  [REEMPLAZANDO] {filepath.name} — reseteando archivo existente...")
            if not resetear_archivo_contenido(session, guid):
                raise RuntimeError("Error al resetear archivo del contenido existente")

            # 2. Obtener nuevo endpoint de upload y subir
            upload_info = obtener_upload_info(session, guid)
            ok_upload = subir_archivo(
                session, upload_info["endpoint"], upload_info["token"], filepath
            )
            if not ok_upload:
                raise RuntimeError("El endpoint de re-upload no devolvió success")

            esperar_procesado(session, guid)

            # 3. Actualizar metadata (merge de niveles/años/etc.)
            ok_meta = actualizar_metadata(
                session, guid,
                nombre,
                erp_id,
                type_g,
                item["is_teacher_only"],
                education_levels=lvls,
                education_years=yrs,
                disciplines=disc,
                langs=langs_act,
                collections=cols,
            )
            if not ok_meta:
                raise RuntimeError("Error al actualizar metadatos tras re-upload")
            resultado["estado"]     = "REEMPLAZADO"
            resultado["url_viewer"] = viewer_url(guid)

        else:
            # ── Contenido nuevo: subir archivo y guardar metadata ────
            resultado["guid"] = guid
            upload_info = obtener_upload_info(session, guid)

            ok_upload = subir_archivo(
                session, upload_info["endpoint"], upload_info["token"], filepath
            )
            if not ok_upload:
                raise RuntimeError("El endpoint de upload no devolvió success")

            esperar_procesado(session, guid)

            ok_meta = actualizar_metadata(
                session, guid, nombre, erp_id, type_g, item["is_teacher_only"],
                education_levels=[level_guid],
                education_years=[year_guid],
                disciplines=[disc_guid],
                langs=[idioma_val],
                collections=[col_guid],
            )
            if not ok_meta:
                raise RuntimeError("Error al actualizar metadatos (PUT)")
            resultado["estado"]     = "OK"
            resultado["url_viewer"] = viewer_url(guid)

    except Exception as e:
        resultado["error"] = str(e)
    finally:
        if is_zip:
            _zip_sem.release()
    return resultado


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  Subidor de contenidos — Compartir Conocimientos Santillana")
    print("=" * 65)

    # ── 1. Ruta del proyecto ────────────────────────────────────────
    # Uso:  python subir_contenidos.py "C:\...\CDCOMPRI1P"
    # Si no se pasa argumento, se solicita de forma interactiva.
    if len(sys.argv) > 1:
        raiz = Path(sys.argv[1].strip('"').strip("'"))
    else:
        print("\nIngresa la ruta de la carpeta raíz del proyecto")
        print("Ejemplo: C:\\Users\\bangulo\\...\\CDCOMPRI1P")
        ruta_str = input("\nRuta > ").strip().strip('"').strip("'")
        if not ruta_str:
            sys.exit("No se ingresó ninguna ruta.")
        raiz = Path(ruta_str)

    output_dir   = raiz / "OUTPUT"
    completo_dir = output_dir / "Completo"

    print(f"\n  Directorio raíz   : {raiz}")
    print(f"  Carpeta OUTPUT    : {output_dir}")
    print(f"  Carpeta Completo  : {completo_dir}")

    if not raiz.is_dir():
        sys.exit(f"\n[ERROR] La carpeta no existe: {raiz}")
    if not output_dir.is_dir():
        sys.exit(f"\n[ERROR] No se encontró OUTPUT/ en {raiz}")
    if not completo_dir.is_dir():
        sys.exit(f"\n[ERROR] No se encontró OUTPUT/Completo/ en {raiz}")

    # ── 2. Token JWT ────────────────────────────────────────────────
    print("\nPega tu token JWT (eyJ...).")
    token = input("Token > ").strip()
    if not token:
        sys.exit("No se ingresó token.")
    session = build_session(token)

    # Buscar Excel en OUTPUT/
    excels = list(output_dir.glob("*.xlsx"))
    if not excels:
        sys.exit(f"\n[ERROR] No se encontró ningún .xlsx en {output_dir}")
    if len(excels) == 1:
        xlsx_path = excels[0]
        print(f"  Excel detectado   : {xlsx_path.name}")
    else:
        print("\nVarios Excel encontrados en OUTPUT/:")
        for i, p in enumerate(excels, 1):
            print(f"  {i}. {p.name}")
        idx = int(input("Elige número > ").strip()) - 1
        xlsx_path = excels[idx]

    # ── 3. Escanear Completo/ y cruzar con Excel ────────────────────
    # Fuente de verdad = archivos físicos en Completo/
    # El Excel aporta: nombre_visible y carpeta_fuente
    print(f"\nLeyendo {xlsx_path.name}...")
    meta_excel = {e["nombre_archivo"]: e for e in leer_excel(xlsx_path)}
    print(f"  {len(meta_excel)} entradas en el Excel.")

    EXTENSIONES_VALIDAS = set(TYPE_GUID.keys())
    archivos_fisicos = sorted(
        f for f in completo_dir.iterdir()
        if f.is_file() and f.suffix.lower() in EXTENSIONES_VALIDAS
    )
    print(f"  {len(archivos_fisicos)} archivos encontrados en Completo/.")

    sin_excel = [f.name for f in archivos_fisicos if f.name not in meta_excel]
    if sin_excel:
        print(f"\n  [INFO] {len(sin_excel)} archivo(s) en Completo/ sin entrada en el Excel")
        print(f"  (se subirán usando el nombre del archivo como nombre visible):")
        for n in sin_excel[:10]:
            print(f"    - {n}")
        if len(sin_excel) > 10:
            print(f"    ... y {len(sin_excel)-10} más")

    # Construir lista final desde los archivos físicos
    archivos = []
    for filepath in archivos_fisicos:
        ext      = filepath.suffix.lower()
        erp_id   = filepath.stem
        type_g   = TYPE_GUID[ext]
        meta     = meta_excel.get(filepath.name, {})
        archivos.append({
            "nombre_visible":   meta.get("nombre_visible", erp_id),
            "nombre_archivo":   filepath.name,
            "erp_id":           erp_id,
            "type_guid":        type_g,
            "carpeta":          meta.get("carpeta", ""),
            "tipo":             meta.get("tipo", "ARCHIVO"),
            "carpeta_fuente":   meta.get("carpeta_fuente", "sin_categoria"),
            "is_teacher_excel": meta.get("is_teacher_excel"),  # None si columna G vacía
            "ruta":             filepath,
        })
    print(f"  Total a subir: {len(archivos)} archivos.\n")

    # ── 4. Selección de metadatos ────────────────────────────────────
    print("\n" + "=" * 65)
    print("  Define los metadatos del lote completo")
    print("=" * 65)

    # Etapa / Education Level
    print("\nObteniendo etapas educativas...")
    levels = fetch_education_levels(session)
    level_guid = menu("Etapa educativa", levels)

    # Año / Education Year
    print("\nObteniendo años/series...")
    years = fetch_education_years(session, level_guid)
    if years:
        year_guid = menu("Año / Serie", years)
    else:
        print("  No se pudo obtener la lista de años. Ingresa el GUID manualmente.")
        year_guid = input("GUID año > ").strip()

    # Asignatura / Discipline
    print("\nObteniendo asignaturas...")
    disciplines = fetch_disciplines(session)
    if disciplines:
        disc_guid = menu("Asignatura / Disciplina", disciplines)
    else:
        print("  No se pudo obtener la lista. Ingresa el GUID manualmente.")
        disc_guid = input("GUID asignatura > ").strip()

    # Idioma
    idioma_val, idioma_label = menu_simple(
        "Idioma",
        {"1": ("es", "Español"), "2": ("en", "Inglés"), "3": ("pt", "Portugués")}
    )

    # Colección
    print("\nBuscando colecciones (escribe parte del nombre para filtrar):")
    busq = input("Búsqueda colección > ").strip()
    collections_list = fetch_collections(session, search=busq)
    if not collections_list:
        print("  Sin resultados. Ingresa el GUID de la colección manualmente.")
        col_guid = input("GUID colección > ").strip()
        col_label = col_guid
    else:
        col_guid = menu("Colección", collections_list)
        col_label = next((c["name"] for c in collections_list if c["guid"] == col_guid), col_guid)

    # ── 5. Disponible para — tabla interactiva por archivo ──────────
    TYPE_LABEL = {
        "CTTY_05": "PDF",
        "CTTY_08": "HTML Interactivo",
        "CTTY_12": "Office",
    }
    DISP_LABEL = {0: "Docentes y Estudiantes", 1: "Solo Docentes"}

    def default_disp(nombre: str) -> int:
        # ZIP con patrón de unidad (U01, U1 ... U08) → ambos; resto → solo docentes
        return 0 if (nombre.lower().endswith(".zip")
                     and re.search(r'[Uu]\d{1,2}(?!\d)', nombre)) else 1

    for a in archivos:
        # Columna G del Excel tiene prioridad; si no existe, usar regex
        if a.get("is_teacher_excel") is not None:
            a["is_teacher_only"] = a["is_teacher_excel"]
        else:
            a["is_teacher_only"] = default_disp(a["nombre_archivo"])

    def _tabla_disp():
        print("\n" + "=" * 88)
        print("  DISPONIBLE PARA — revisa y edita (Enter vacío para confirmar todo)")
        print("  Fuente: [X]=columna G del Excel  [ ]=regla automática (ZIP+unidad / resto)")
        print("=" * 88)
        print(f"  {'#':<5} {'Archivo':<38} {'Tipo':<18} {'Disponible para':<25} {'Fuente'}")
        print("  " + "─" * 84)
        for i, a in enumerate(archivos, 1):
            tipo_lbl   = TYPE_LABEL.get(a["type_guid"], a["type_guid"])
            disp_lbl   = DISP_LABEL[a["is_teacher_only"]]
            fuente_lbl = "[X] Excel" if a.get("is_teacher_excel") is not None else "[ ] auto"
            print(f"  {i:<5} {a['nombre_archivo'][:37]:<38} {tipo_lbl:<18} {disp_lbl:<25} {fuente_lbl}")
        print("  " + "─" * 84)

    while True:
        _tabla_disp()
        entrada = input("\n  Nro a editar (Enter para confirmar todos) > ").strip()
        if entrada == "":
            break
        if entrada.isdigit() and 1 <= int(entrada) <= len(archivos):
            idx = int(entrada) - 1
            a = archivos[idx]
            print(f"\n  Archivo : {a['nombre_archivo']}")
            print(f"  Actual  : {DISP_LABEL[a['is_teacher_only']]}")
            print("    1. Docentes y Estudiantes")
            print("    2. Solo Docentes")
            resp = input("  Nuevo valor (1/2) > ").strip()
            if resp == "1":
                a["is_teacher_only"] = 0
                print("  → Docentes y Estudiantes")
            elif resp == "2":
                a["is_teacher_only"] = 1
                print("  → Solo Docentes")
            else:
                print("  Sin cambios.")
        else:
            print(f"  Número no válido (1–{len(archivos)}).")

    # ── 6. Previsualización detallada ────────────────────────────────
    level_name = next((l["name"] for l in levels if l["guid"] == level_guid), level_guid)
    year_name  = next((y["name"] for y in years  if y["guid"] == year_guid),  year_guid) if years else year_guid
    disc_name  = next((d["name"] for d in disciplines if d["guid"] == disc_guid), disc_guid) if disciplines else disc_guid

    print("\n" + "=" * 100)
    print("  DETALLE COMPLETO — así se subirá cada archivo")
    print("=" * 100)
    print(f"  {'#':<4} {'Nombre visible':<38} {'Archivo':<36} {'Tipo':<18} {'Disponible para'}")
    print("  " + "─" * 96)
    for i, a in enumerate(archivos, 1):
        tipo_label = TYPE_LABEL.get(a["type_guid"], a["type_guid"])
        disp_label = DISP_LABEL[a["is_teacher_only"]]
        nombre_vis = a["nombre_visible"][:37]
        nombre_arc = a["nombre_archivo"][:35]
        print(f"  {i:<4} {nombre_vis:<38} {nombre_arc:<36} {tipo_label:<18} {disp_label}")
    print("=" * 100)

    print(f"\n  METADATOS COMUNES A TODOS:")
    print(f"    Etapa           : {level_name}")
    print(f"    Año / Serie     : {year_name}")
    print(f"    Asignatura      : {disc_name}")
    print(f"    Idioma          : {idioma_label}")
    print(f"    Colección       : {col_label}")
    print(f"\n  Total: {len(archivos)} archivos")
    print("=" * 100)

    conf = input("\n¿Iniciar subida? (s/n) > ").strip().lower()
    if conf != "s":
        sys.exit("Cancelado por el usuario.")

    # ── 7. Subida en paralelo (pool de MAX_WORKERS) ──────────────────
    print(f"\n  Subiendo {len(archivos)} archivos — máximo {MAX_WORKERS} simultáneos\n")
    resultados = []
    barra = tqdm(total=len(archivos), desc="Completados", unit="arch", ncols=80)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(
                _subir_item, token,
                item, level_guid, year_guid, disc_guid, idioma_val, col_guid
            ): item
            for item in archivos
        }
        for future in as_completed(futures):
            resultado = future.result()
            resultados.append(resultado)
            if resultado["estado"] == "OK":
                tqdm.write(f"  [OK]          {resultado['archivo']}")
            elif resultado["estado"] == "REEMPLAZADO":
                tqdm.write(f"  [REEMPLAZADO] {resultado['archivo']}")
            elif resultado["estado"] == "ACTUALIZADO":
                tqdm.write(f"  [ACTUALIZADO] {resultado['archivo']}")
            else:
                tqdm.write(f"  [ERROR]       {resultado['archivo']}: {resultado['error']}")
            barra.update(1)

    barra.close()

    # ── 7. Log de resultados ─────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = raiz / f"log_subida_{ts}.xlsx"
    guardar_log(resultados, log_path)

    ok_count   = sum(1 for r in resultados if r["estado"] == "OK")
    rep_count  = sum(1 for r in resultados if r["estado"] == "REEMPLAZADO")
    upd_count  = sum(1 for r in resultados if r["estado"] == "ACTUALIZADO")
    err_count  = len(resultados) - ok_count - rep_count - upd_count

    print(f"\n{'=' * 65}")
    print(f"  Subida completada.")
    print(f"  Nuevos       : {ok_count}")
    print(f"  Reemplazados : {rep_count}")
    print(f"  Actualizados : {upd_count}")
    print(f"  Errores      : {err_count}")
    print(f"  Log          : {log_path}")
    print("=" * 65)


if __name__ == "__main__":
    main()
