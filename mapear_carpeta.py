#!/usr/bin/env python3
"""
Mapea la estructura de una carpeta local.
Genera:
  - mapa_carpeta.txt     → árbol visual de carpetas y archivos
  - resumen_carpeta.xlsx → tabla con Carpeta / Archivo / Extensión / Tamaño / Ruta
Uso:
  python mapear_carpeta.py "C:\\..\\CDCOMPRI1P"
  python mapear_carpeta.py          (pide la ruta de forma interactiva)
"""

import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Instalando dependencias necesarias...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment


# ── Árbol ────────────────────────────────────────────────────────────────────

TEE   = "├── "
LAST  = "└── "
VERT  = "│   "
BLANK = "    "


def build_tree(path: Path, prefix: str = "") -> list[str]:
    """Genera líneas de árbol para una carpeta, recursivamente."""
    lines = []
    entries = sorted(path.iterdir(), key=lambda p: (p.is_file(), p.name.lower()))
    for i, entry in enumerate(entries):
        is_last = (i == len(entries) - 1)
        connector = LAST if is_last else TEE
        lines.append(f"{prefix}{connector}{entry.name}{'/' if entry.is_dir() else ''}")
        if entry.is_dir():
            ext_prefix = BLANK if is_last else VERT
            lines.extend(build_tree(entry, prefix + ext_prefix))
    return lines


# ── Escaneo de archivos ──────────────────────────────────────────────────────

def human_size(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024:
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} TB"


def scan_files(raiz: Path) -> list[dict]:
    """Escanea recursivamente y devuelve lista de dicts con info de cada archivo."""
    archivos = []
    for f in sorted(raiz.rglob("*")):
        if f.is_file():
            rel     = f.relative_to(raiz)
            carpeta = str(rel.parent) if str(rel.parent) != "." else "(raíz)"
            archivos.append({
                "carpeta":   carpeta,
                "nombre":    f.name,
                "extension": f.suffix.lower() or "(sin ext)",
                "tamaño":    f.stat().st_size,
                "ruta":      str(f.resolve()),
            })
    return archivos


# ── Excel ────────────────────────────────────────────────────────────────────

def crear_excel(archivos: list[dict], ruta_xlsx: Path):
    wb = openpyxl.Workbook()

    fill_h = PatternFill("solid", fgColor="1F4E79")
    font_h = Font(bold=True, color="FFFFFF", size=11)
    fill_par   = PatternFill("solid", fgColor="DDEEFF")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")

    # ── Hoja Archivos ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Archivos"
    headers = ["Carpeta", "Archivo", "Extensión", "Tamaño", "Ruta completa"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_h
        c.font = font_h
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    for i, a in enumerate(archivos, 2):
        fill = fill_par if i % 2 == 0 else fill_impar
        for col, val in enumerate(
            [a["carpeta"], a["nombre"], a["extension"], human_size(a["tamaño"]), a["ruta"]], 1
        ):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill
            c.alignment = Alignment(wrap_text=(col == 5))

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 80

    # ── Hoja Por extensión ──────────────────────────────────────────
    ws2 = wb.create_sheet("Por extensión")
    ws2.append(["Extensión", "Archivos", "Tamaño total"])
    for col in ("A", "B", "C"):
        ws2[f"{col}1"].fill = fill_h
        ws2[f"{col}1"].font = font_h

    conteo: dict[str, dict] = defaultdict(lambda: {"archivos": 0, "bytes": 0})
    for a in archivos:
        conteo[a["extension"]]["archivos"] += 1
        conteo[a["extension"]]["bytes"]    += a["tamaño"]

    for ext, datos in sorted(conteo.items(), key=lambda x: -x[1]["archivos"]):
        ws2.append([ext, datos["archivos"], human_size(datos["bytes"])])

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 15

    wb.save(ruta_xlsx)
    print(f"  Guardado: {ruta_xlsx}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  Mapeador de carpeta local")
    print("=" * 62)

    if len(sys.argv) > 1:
        raiz = Path(sys.argv[1].strip('"').strip("'"))
    else:
        print("\nIngresa la ruta de la carpeta a mapear.")
        print("Ejemplo: C:\\Users\\bangulo\\...\\CDCOMPRI1P")
        ruta_str = input("\nRuta > ").strip().strip('"').strip("'")
        if not ruta_str:
            sys.exit("No se ingresó ninguna ruta.")
        raiz = Path(ruta_str)

    if not raiz.is_dir():
        sys.exit(f"\n[ERROR] La carpeta no existe: {raiz}")

    TXT_OUT  = raiz / "mapa_carpeta.txt"
    XLSX_OUT = raiz / "resumen_carpeta.xlsx"

    print(f"\n  Carpeta raíz : {raiz.resolve()}")

    # ── Árbol TXT ────────────────────────────────────────────────────
    print(f"\n[1/2] Generando árbol → {TXT_OUT.name}")
    lineas = [f"{raiz.name}/"] + build_tree(raiz)
    TXT_OUT.write_text("\n".join(lineas), encoding="utf-8")

    print()
    print("─" * 62)
    for l in lineas:
        print(l)
    print("─" * 62)

    # ── Escaneo y Excel ──────────────────────────────────────────────
    print(f"\n[2/2] Escaneando archivos y generando Excel → {XLSX_OUT.name}")
    archivos = scan_files(raiz)
    crear_excel(archivos, XLSX_OUT)

    # ── Estadísticas ─────────────────────────────────────────────────
    total_bytes = sum(a["tamaño"] for a in archivos)
    conteo_ext  = Counter(a["extension"] for a in archivos)

    print(f"\n{'=' * 62}")
    print(f"  Archivos  : {len(archivos)}")
    print(f"  Tamaño    : {human_size(total_bytes)}")
    print()
    for ext, n in conteo_ext.most_common():
        print(f"    {ext:<14} {n} archivo{'s' if n != 1 else ''}")
    print()
    print(f"  Árbol     : {TXT_OUT.resolve()}")
    print(f"  Excel     : {XLSX_OUT.resolve()}")
    print("=" * 62)


if __name__ == "__main__":
    main()
