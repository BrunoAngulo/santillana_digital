#!/usr/bin/env python3
"""
Mapea la estructura de documentos de Santillana Digital sin descargar nada.
Genera:
  - mapa_estructura.txt  → árbol visual de carpetas y archivos
  - resumen_global.xlsx  → tabla con Producto / Carpeta / Archivo / URL
"""

import re
import sys
import time
from collections import defaultdict
from pathlib import Path

try:
    import requests
    from bs4 import BeautifulSoup
    from tqdm import tqdm
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Instalando dependencias necesarias...")
    import subprocess
    subprocess.check_call([
        sys.executable, "-m", "pip", "install",
        "requests", "beautifulsoup4", "tqdm", "openpyxl",
    ])
    import requests
    from bs4 import BeautifulSoup
    from tqdm import tqdm
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment


BASE_URL  = "https://digital.santillana.com.pe"
DOCS_URL  = f"{BASE_URL}/Documentos/SantiVaContigoDocs"
TXT_OUT   = Path("mapa_estructura.txt")
XLSX_OUT  = Path("resumen_global.xlsx")
ROOT_LABEL = "SantillanaDigital_Docs"


# ── Utilidades ──────────────────────────────────────────────────────────────

def sanitize(name: str) -> str:
    name = name.strip()
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    name = re.sub(r'\s+', ' ', name)
    name = name.rstrip('.')
    return name or "sin_nombre"


def build_session(cookie_str: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Referer": BASE_URL,
        "Cookie": cookie_str.strip(),
    })
    return s


def get_soup(session, url: str):
    for intento in range(3):
        try:
            r = session.get(url, timeout=30)
            r.raise_for_status()
            return BeautifulSoup(r.text, "html.parser")
        except requests.RequestException as e:
            if intento < 2:
                time.sleep(2 ** intento)
            else:
                tqdm.write(f"  [ERROR] {url}: {e}")
                return None


def verificar_sesion(session) -> bool:
    try:
        r = session.get(DOCS_URL, timeout=15, allow_redirects=True)
        if "login" in r.url.lower() or "account" in r.url.lower():
            return False
        return "productoID" in r.text or "titulo-contenido" in r.text
    except requests.RequestException:
        return False


# ── Scraping ────────────────────────────────────────────────────────────────

def obtener_productos(session) -> list[dict]:
    soup = get_soup(session, DOCS_URL)
    if not soup:
        return []
    productos = []
    for a in soup.select("a[href*='/Documentos/Contenido']"):
        tag = a.find_next("p", class_="titulo-contenido")
        nombre = tag.get_text(strip=True) if tag else a["href"]
        productos.append({"nombre": nombre, "url": BASE_URL + a["href"]})
    return productos


def obtener_carpetas(session, producto: dict) -> list[dict]:
    soup = get_soup(session, producto["url"])
    if not soup:
        return []
    carpetas = []
    for a in soup.select("a[href*='/Documentos/Archivo']"):
        tag = a.find_next("p", class_="titulo-contenido")
        nombre = tag.get_text(strip=True) if tag else a["href"]
        carpetas.append({"nombre": nombre, "url": BASE_URL + a["href"]})
    return carpetas


def obtener_archivos(session, carpeta: dict) -> list[dict]:
    soup = get_soup(session, carpeta["url"])
    if not soup:
        return []
    archivos = []
    for a in soup.select("a[href]"):
        href = a["href"]
        if "/api/accesoblob/" in href or any(
            href.lower().endswith(ext)
            for ext in (".pdf", ".docx", ".xlsx", ".pptx", ".zip")
        ):
            tag = a.find_next("p", class_="titulo-contenido")
            nombre_display = tag.get_text(strip=True) if tag else ""
            ext = Path(href.split("?")[0]).suffix or ".pdf"
            nombre_archivo = sanitize(nombre_display or Path(href).stem) + ext
            archivos.append({"nombre": nombre_archivo, "url": href})
    return archivos


# ── Generador de árbol ──────────────────────────────────────────────────────

PIPE  = "│"
TEE   = "├── "
LAST  = "└── "
BLANK = "    "
VERT  = "│   "


def tree_lines(estructura: list[dict]) -> list[str]:
    """
    estructura: lista de dicts con keys producto / carpeta / archivos (list[dict])
    Devuelve líneas del árbol listas para unir con \n.
    """
    lines = [f"{ROOT_LABEL}/"]

    # Agrupar: producto → carpeta → archivos
    agrupado: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    for item in estructura:
        agrupado[item["producto"]][item["carpeta"]].extend(item["archivos"])

    # Colección plana para la carpeta Completo
    todos_archivos: list[dict] = []
    for item in estructura:
        todos_archivos.extend(item["archivos"])

    productos = list(agrupado.keys())

    # Añadimos Completo + resumen_global.xlsx como entradas especiales al final
    EXTRA = ["__COMPLETO__", "__XLSX__"]
    total_raiz = len(productos) + len(EXTRA)

    for i, prod in enumerate(productos):
        es_ultimo_prod = (i == total_raiz - len(EXTRA) - 1)
        pref_prod  = LAST if es_ultimo_prod else TEE
        cont_prod  = BLANK if es_ultimo_prod else VERT

        lines.append(f"{pref_prod}{sanitize(prod)}/")

        carpetas = list(agrupado[prod].keys())
        for j, carp in enumerate(carpetas):
            es_ultima_carp = (j == len(carpetas) - 1)
            pref_carp = cont_prod + (LAST if es_ultima_carp else TEE)
            cont_carp = cont_prod + (BLANK if es_ultima_carp else VERT)

            archivos = agrupado[prod][carp]
            lines.append(f"{pref_carp}{sanitize(carp)}/")

            for k, arch in enumerate(archivos):
                es_ultimo_arch = (k == len(archivos) - 1)
                pref_arch = cont_carp + (LAST if es_ultimo_arch else TEE)
                lines.append(f"{pref_arch}{arch['nombre']}")

    # ── Completo/ ──────────────────────────────────────────────────
    lines.append(f"{TEE}Completo/  ← todos los archivos sin subcarpetas")
    for k, arch in enumerate(todos_archivos):
        es_ultimo = (k == len(todos_archivos) - 1)
        pref = VERT + (LAST if es_ultimo else TEE)
        lines.append(f"{pref}{arch['nombre']}")

    # ── resumen_global.xlsx ────────────────────────────────────────
    lines.append(f"{LAST}{XLSX_OUT.name}")

    return lines


# ── Excel ───────────────────────────────────────────────────────────────────

def crear_excel(estructura: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documentos"

    # Cabecera
    headers = ["Producto", "Carpeta", "Archivo", "URL"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 18

    # Filas
    fila = 2
    fill_par   = PatternFill("solid", fgColor="DDEEFF")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")

    for item in estructura:
        for arch in item["archivos"]:
            fill = fill_par if fila % 2 == 0 else fill_impar
            datos = [item["producto"], item["carpeta"], arch["nombre"], arch["url"]]
            for col, val in enumerate(datos, 1):
                cell = ws.cell(row=fila, column=col, value=val)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=(col == 4))
            fila += 1

    # Ancho de columnas
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 80

    # Hoja resumen por producto
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Producto", "Carpetas", "Archivos"])
    ws2["A1"].font = header_font
    ws2["B1"].font = header_font
    ws2["C1"].font = header_font
    for cell in ws2["1:1"]:
        cell.fill = header_fill

    conteo: dict[str, dict] = defaultdict(lambda: {"carpetas": set(), "archivos": 0})
    for item in estructura:
        conteo[item["producto"]]["carpetas"].add(item["carpeta"])
        conteo[item["producto"]]["archivos"] += len(item["archivos"])

    for prod, datos in conteo.items():
        ws2.append([prod, len(datos["carpetas"]), datos["archivos"]])

    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    wb.save(XLSX_OUT)
    print(f"  Guardado: {XLSX_OUT}")


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  Mapeador de Santillana Digital - SantiVaContigoDocs")
    print("=" * 62)
    print()
    print("Pega tu cookie de sesión completa y presiona Enter.")
    print()

    cookie_str = input("Cookie > ").strip()
    if not cookie_str:
        print("No se ingresó ninguna cookie. Saliendo.")
        sys.exit(1)

    print("\nVerificando sesión...", end=" ", flush=True)
    session = build_session(cookie_str)

    if not verificar_sesion(session):
        print("FALLO")
        print("\n[ERROR] La cookie no da acceso. Posibles causas:")
        print("  - La sesión expiró (vuelve a loguearte)")
        print("  - Falta alguna cookie del string")
        sys.exit(1)

    print("OK\n")

    # ── Recolectar estructura ────────────────────────────────────────
    print("Recolectando estructura (sin descargar archivos)...\n")
    productos = obtener_productos(session)
    if not productos:
        print("No se encontraron productos.")
        sys.exit(1)

    print(f"  {len(productos)} productos encontrados.\n")

    estructura: list[dict] = []
    barra = tqdm(productos, desc="Escaneando", unit="prod", ncols=80)

    for prod in barra:
        barra.set_postfix_str(sanitize(prod["nombre"])[:35])
        carpetas = obtener_carpetas(session, prod)
        for carp in carpetas:
            archivos = obtener_archivos(session, carp)
            if archivos:
                estructura.append({
                    "producto": prod["nombre"],
                    "carpeta":  carp["nombre"],
                    "archivos": archivos,
                })

    # ── Generar árbol TXT ────────────────────────────────────────────
    print(f"\n[1/2] Generando árbol de texto → {TXT_OUT}")
    lineas = tree_lines(estructura)
    TXT_OUT.write_text("\n".join(lineas), encoding="utf-8")
    print(f"  {len(lineas)} líneas escritas.")

    # ── Mostrar árbol en consola ─────────────────────────────────────
    print()
    print("─" * 62)
    for l in lineas:
        print(l)
    print("─" * 62)

    # ── Generar Excel ────────────────────────────────────────────────
    print(f"\n[2/2] Generando Excel → {XLSX_OUT}")
    crear_excel(estructura)

    # ── Estadísticas ────────────────────────────────────────────────
    total_arch = sum(len(i["archivos"]) for i in estructura)
    total_carp = len({(i["producto"], i["carpeta"]) for i in estructura})
    print(f"\n{'=' * 62}")
    print(f"  Productos  : {len(productos)}")
    print(f"  Carpetas   : {total_carp}")
    print(f"  Archivos   : {total_arch}")
    print(f"  Árbol      : {TXT_OUT.resolve()}")
    print(f"  Excel      : {XLSX_OUT.resolve()}")
    print("=" * 62)


if __name__ == "__main__":
    main()
